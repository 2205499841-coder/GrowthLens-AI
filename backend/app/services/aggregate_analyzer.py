import hashlib
import logging
import math
import re
from dataclasses import dataclass
from io import BytesIO
from statistics import median
from typing import Any, Callable, Iterable

import pandas as pd
from openpyxl import load_workbook

from app.schemas.aggregate_analysis import AggregateAnalysisResponse
from app.services.schema_mapper import ExtractedExcelSchema


logger = logging.getLogger(__name__)


TOTAL_ROW_LABELS = {
    "总计",
    "合计",
    "汇总",
    "全部",
    "整体",
    "overall",
    "total",
}
GROUP_HEADER_HINTS = (
    "用户规模",
    "流量指标",
    "转化指标",
    "经营指标",
    "本期",
    "同比",
    "环比",
)
FILTER_HINTS = (
    "统计周期",
    "报表周期",
    "数据周期",
    "时间范围",
    "渠道筛选",
    "用户类型",
    "地区筛选",
)
COMPARISON_HINTS = (
    "同比",
    "同期",
    "环比",
    "环期",
    "上期",
    "较上期",
    "百分点",
)
COMPARISON_DELTA_HINTS = (
    "偏差",
    "变化",
    "增长",
    "下降",
    "增幅",
    "差异",
    "增减",
)


@dataclass(frozen=True)
class _MetricRule:
    metric_key: str
    label: str
    aliases: tuple[str, ...]
    role: str
    unit: str
    aggregation: str
    stage_order: int | None = None
    semantic_role: str = "outcome_metric"


METRIC_RULES = (
    _MetricRule(
        "all_channel_payment_users",
        "全域支付用户",
        ("全域支付用户数", "全域支付人数", "全渠道支付用户数"),
        "count_metric",
        "count",
        "non_additive",
    ),
    _MetricRule(
        "online_payment_users",
        "线上支付用户",
        ("线上支付用户数", "线上支付人数", "线上支付用户"),
        "count_metric",
        "count",
        "non_additive",
    ),
    _MetricRule(
        "store_payment_users",
        "门店支付用户",
        (
            "门店签到支付用户数",
            "门店签到支付用户",
            "门店支付用户数",
            "到店支付用户数",
        ),
        "count_metric",
        "count",
        "non_additive",
    ),
    _MetricRule(
        "traffic_users",
        "浏览用户",
        (
            "浏览用户数",
            "浏览人数",
            "访问用户数",
            "访问人数",
            "访客数",
            "访客人数",
            "流量用户数",
            "uv",
        ),
        "count_metric",
        "count",
        "non_additive",
        10,
        "funnel_stage",
    ),
    _MetricRule(
        "product_detail_users",
        "商详用户",
        (
            "商详用户数",
            "商详人数",
            "商品详情用户数",
            "详情页用户数",
        ),
        "count_metric",
        "count",
        "non_additive",
        20,
        "funnel_stage",
    ),
    _MetricRule(
        "lead_users",
        "咨询用户",
        ("咨询用户数", "留资用户数", "线索用户数"),
        "count_metric",
        "count",
        "non_additive",
        25,
        "funnel_stage",
    ),
    _MetricRule(
        "appointment_users",
        "预约用户",
        ("预约用户数", "预约人数", "发起预约用户数"),
        "count_metric",
        "count",
        "non_additive",
        30,
        "funnel_stage",
    ),
    _MetricRule(
        "sku_selection_users",
        "SKU 选择用户",
        ("sku选择用户数", "选择sku用户数", "规格选择用户数", "选规格用户数"),
        "count_metric",
        "count",
        "non_additive",
        40,
        "funnel_stage",
    ),
    _MetricRule(
        "time_confirmation_users",
        "时间确认用户",
        ("预约时间确认用户数", "时间确认用户数", "档期确认用户数"),
        "count_metric",
        "count",
        "non_additive",
        50,
        "funnel_stage",
    ),
    _MetricRule(
        "order_submission_users",
        "提交订单用户",
        ("提交订单用户数", "订单提交用户数", "下单用户数", "下单人数"),
        "count_metric",
        "count",
        "non_additive",
        60,
        "funnel_stage",
    ),
    _MetricRule(
        "visit_users",
        "到店用户",
        ("到店用户数", "到店人数", "核销用户数"),
        "count_metric",
        "count",
        "non_additive",
        70,
        "funnel_stage",
    ),
    _MetricRule(
        "payment_users",
        "支付用户",
        (
            "支付用户数",
            "支付人数",
            "成交用户数",
            "成交人数",
        ),
        "count_metric",
        "count",
        "non_additive",
        80,
        "funnel_stage",
    ),
    _MetricRule(
        "traffic_to_detail_rate",
        "浏览至商详转化率",
        (
            "浏览至商详转化率",
            "浏览到商详转化率",
            "浏览→商详转化率",
            "浏览至商详",
            "浏览到商详",
            "浏览→商详",
            "商详转化率",
        ),
        "rate_metric",
        "ratio",
        "weighted_rate",
    ),
    _MetricRule(
        "detail_to_appointment_rate",
        "商详至预约转化率",
        (
            "商详至预约转化率",
            "商详到预约转化率",
            "商详→预约转化率",
            "商详至预约",
            "商详到预约",
            "商详→预约",
        ),
        "rate_metric",
        "ratio",
        "weighted_rate",
    ),
    _MetricRule(
        "appointment_to_sku_rate",
        "预约至 SKU 选择转化率",
        (
            "预约至sku转化率",
            "预约到sku转化率",
            "预约→sku转化率",
            "预约至sku",
            "预约到sku",
            "预约→sku",
        ),
        "rate_metric",
        "ratio",
        "weighted_rate",
    ),
    _MetricRule(
        "sku_to_time_confirmation_rate",
        "SKU 至时间确认转化率",
        (
            "sku至时间确认转化率",
            "sku到时间确认转化率",
            "sku→时间确认转化率",
            "sku至时间确认",
            "sku到时间确认",
            "sku→时间确认",
        ),
        "rate_metric",
        "ratio",
        "weighted_rate",
    ),
    _MetricRule(
        "time_confirmation_to_order_rate",
        "时间确认至提交订单转化率",
        (
            "时间确认至提交订单转化率",
            "时间确认到下单转化率",
            "时间确认→提交订单转化率",
            "时间确认至提交订单",
            "时间确认到下单",
            "时间确认→提交订单",
        ),
        "rate_metric",
        "ratio",
        "weighted_rate",
    ),
    _MetricRule(
        "order_to_payment_rate",
        "提交订单至支付转化率",
        (
            "提交订单至支付转化率",
            "下单到支付转化率",
            "提交订单→支付转化率",
            "提交订单至支付",
            "下单到支付",
            "提交订单→支付",
        ),
        "rate_metric",
        "ratio",
        "weighted_rate",
    ),
    _MetricRule(
        "payment_conversion_rate",
        "支付转化率",
        (
            "浏览至支付转化率",
            "浏览到支付转化率",
            "浏览→支付转化率",
            "支付转化率",
            "成交转化率",
            "成交率",
        ),
        "rate_metric",
        "ratio",
        "weighted_rate",
    ),
    _MetricRule(
        "average_order_value",
        "客单价",
        ("客单价", "平均订单金额", "平均支付金额"),
        "amount_metric",
        "currency_per_order",
        "non_additive",
    ),
    _MetricRule(
        "gmv",
        "GMV",
        ("gmv", "交易额", "成交金额", "销售额", "支付金额"),
        "amount_metric",
        "currency",
        "sum",
    ),
)

DIMENSION_RULES = (
    ("category", "品类", ("品类", "类目", "分类", "业务类型")),
    ("channel", "渠道", ("渠道", "来源渠道", "获客渠道")),
    ("region", "地区", ("地区", "区域", "城市", "省份")),
    ("user_type", "用户类型", ("用户类型", "客群", "用户分层", "人群")),
    ("page_version", "页面版本", ("页面版本", "版本", "页面类型")),
)

KPI_ORDER = (
    "traffic_users",
    "appointment_users",
    "payment_users",
    "payment_conversion_rate",
    "gmv",
    "average_order_value",
)

FUNNEL_RATE_KEYS = {
    ("traffic_users", "product_detail_users"): "traffic_to_detail_rate",
    ("product_detail_users", "appointment_users"): (
        "detail_to_appointment_rate"
    ),
    ("appointment_users", "sku_selection_users"): (
        "appointment_to_sku_rate"
    ),
    ("sku_selection_users", "time_confirmation_users"): (
        "sku_to_time_confirmation_rate"
    ),
    ("time_confirmation_users", "order_submission_users"): (
        "time_confirmation_to_order_rate"
    ),
    ("order_submission_users", "payment_users"): "order_to_payment_rate",
}


@dataclass(frozen=True)
class _ColumnProfile:
    source_column: str
    column_index: int
    inferred_type: str
    numeric_ratio: float
    unique_ratio: float
    percentage_format: bool
    value_range: str

    def ai_summary(self) -> dict[str, Any]:
        return {
            "source_column": self.source_column,
            "inferred_type": self.inferred_type,
            "numeric_ratio": round(self.numeric_ratio, 2),
            "unique_ratio": round(self.unique_ratio, 2),
            "percentage_format": self.percentage_format,
            "value_range": self.value_range,
        }


@dataclass(frozen=True)
class _SemanticField:
    source_column: str
    label: str
    semantic_key: str
    role: str
    unit: str
    aggregation: str
    confidence: str
    stage_order: int | None = None
    semantic_role: str = "outcome_metric"
    comparison_type: str | None = None
    comparison_period: str | None = None
    comparison_unit: str | None = None
    comparison_value_kind: str | None = None
    target_metric_key: str | None = None


@dataclass(frozen=True)
class _ParsedAggregateWorkbook:
    sheet_name: str
    header_rows: tuple[int, ...]
    data_frame: pd.DataFrame
    number_formats: dict[str, tuple[str, ...]]
    report_period: str | None
    filters: dict[str, str]
    warnings: tuple[str, ...]


AggregateFallbackResolver = Callable[
    [list[dict[str, Any]]],
    list[dict[str, Any]],
]


def analyze_aggregate_excel(
    file_content: bytes,
    *,
    file_name: str,
    extracted_schema: ExtractedExcelSchema,
    fallback_resolver: AggregateFallbackResolver | None = None,
) -> AggregateAnalysisResponse:
    parsed = _parse_aggregate_workbook(file_content, extracted_schema)
    profiles = _build_column_profiles(
        parsed.data_frame,
        parsed.number_formats,
    )
    semantic_fields, unrecognized_columns = _recognize_semantic_fields(
        profiles,
        fallback_resolver=fallback_resolver,
    )
    dimensions = [
        field for field in semantic_fields if field.role == "dimension"
    ]
    metrics = [
        field for field in semantic_fields if field.role != "dimension"
    ]
    business_metrics = [
        field for field in metrics if field.role != "comparison_metric"
    ]
    comparisons = [
        field for field in metrics if field.role == "comparison_metric"
    ]
    funnel_definitions = _build_funnel_definitions(business_metrics)

    primary_dimension = dimensions[0] if dimensions else None
    total_rows, detail_rows = _split_total_and_detail_rows(
        parsed.data_frame,
        primary_dimension,
    )
    warnings = list(parsed.warnings)
    if len(total_rows) > 1:
        warnings.append("检测到多个总计行，整体指标使用首个总计行。")
    if not total_rows and len(detail_rows) > 1:
        warnings.append(
            "未检测到总计行；用户数和转化率不跨维度汇总。"
        )

    scope_row, scope_source = _select_scope_row(total_rows, detail_rows)
    metric_lookup = _deduplicate_metrics(business_metrics)
    comparison_lookup = _group_comparisons(comparisons)
    kpis = _build_kpis(
        scope_row,
        scope_source,
        detail_rows,
        metric_lookup,
        parsed.number_formats,
    )
    dimension_performance = _build_dimension_performance(
        detail_rows,
        primary_dimension,
        metric_lookup,
        comparison_lookup,
        parsed.number_formats,
    )
    dimension_funnel_diagnostics = _build_dimension_funnel_diagnostics(
        detail_rows,
        primary_dimension,
        funnel_definitions,
        metric_lookup,
        comparison_lookup,
        parsed.number_formats,
        dimension_performance,
    )
    _apply_final_comparisons_to_performance(
        dimension_performance,
        dimension_funnel_diagnostics,
    )
    funnel, funnel_warnings = _build_dynamic_funnel(
        scope_row,
        scope_source,
        funnel_definitions,
        comparison_lookup,
        parsed.number_formats,
        primary_dimension,
    )
    warnings.extend(funnel_warnings)
    diagnostics = _build_diagnostics(
        dimension_performance,
        dimension_funnel_diagnostics,
        detail_rows,
        primary_dimension,
        metric_lookup,
        comparison_lookup,
        parsed.number_formats,
    )
    opportunities = _build_opportunities(dimension_performance)
    business_insights = _build_business_insights(
        dimension_funnel_diagnostics,
        diagnostics,
        opportunities,
        dimension_performance,
    )

    recognized_column_count = len(
        {field.source_column for field in semantic_fields}
    )
    meaningful_column_count = max(len(profiles), 1)
    recognition_ratio = recognized_column_count / meaningful_column_count
    analysis_status = _resolve_analysis_status(
        dimensions=dimensions,
        business_metrics=business_metrics,
        detail_rows=detail_rows,
        recognition_ratio=recognition_ratio,
        warnings=warnings,
    )

    return AggregateAnalysisResponse.model_validate(
        {
            "dataset_type": "aggregate_metrics",
            "analysis_status": analysis_status,
            "metadata": {
                "file_name": file_name,
                "data_start_date": None,
                "data_end_date": None,
            },
            "dataset": {
                "dataset_type": "aggregate_metrics",
                "analysis_status": analysis_status,
                "sheet_name": parsed.sheet_name,
                "header_rows": list(parsed.header_rows),
                "grain": [field.semantic_key for field in dimensions],
                "report_period": parsed.report_period,
                "filters": parsed.filters,
            },
            "dimensions": [
                {
                    "source_column": field.source_column,
                    "label": field.label,
                    "semantic_key": field.semantic_key,
                    "confidence": field.confidence,
                }
                for field in dimensions
            ],
            "metrics": [
                {
                    "source_column": field.source_column,
                    "label": field.label,
                    "metric_key": field.semantic_key,
                    "role": field.role,
                    "unit": field.unit,
                    "aggregation": field.aggregation,
                    "confidence": field.confidence,
                    "semantic_role": field.semantic_role,
                }
                for field in metrics
            ],
            "funnel_stages": [
                {
                    "metric_key": field.semantic_key,
                    "label": field.label,
                    "stage_order": field.stage_order,
                    "source_column": field.source_column,
                    "confidence": field.confidence,
                }
                for field in funnel_definitions
            ],
            "comparisons": [
                {
                    "source_column": field.source_column,
                    "label": field.label,
                    "comparison_type": field.comparison_type,
                    "period": field.comparison_period,
                    "unit": field.comparison_unit,
                    "value_kind": field.comparison_value_kind,
                    "target_metric_key": field.target_metric_key,
                    "confidence": field.confidence,
                }
                for field in comparisons
            ],
            "data_quality": {
                "row_count": len(parsed.data_frame),
                "detail_row_count": len(detail_rows),
                "total_row_detected": bool(total_rows),
                "recognized_column_count": recognized_column_count,
                "unrecognized_columns": unrecognized_columns,
                "warnings": list(dict.fromkeys(warnings)),
            },
            "kpis": kpis,
            "funnel": funnel,
            "dimension_performance": dimension_performance,
            "dimension_funnel_diagnostics": dimension_funnel_diagnostics,
            "diagnostics": diagnostics,
            "opportunities": opportunities,
            "business_insights": business_insights,
        }
    )


def _parse_aggregate_workbook(
    file_content: bytes,
    extracted_schema: ExtractedExcelSchema,
) -> _ParsedAggregateWorkbook:
    workbook = load_workbook(BytesIO(file_content), data_only=True)
    try:
        worksheet = workbook[extracted_schema.sheet_name]
        max_column = min(max(worksheet.max_column, 1), 200)
        header_row = extracted_schema.header_row_index
        header_rows = _resolve_header_rows(
            worksheet,
            header_row,
            max_column,
        )
        expanded_headers = [
            _expanded_row_values(worksheet, row_index, max_column)
            for row_index in header_rows
        ]
        headers = _flatten_headers(expanded_headers)
        active_columns = [
            index for index, header in enumerate(headers, start=1) if header
        ]
        rows: list[list[Any]] = []
        warnings: list[str] = []
        number_formats: dict[str, list[str]] = {
            headers[index - 1]: [] for index in active_columns
        }

        for row_index in range(header_rows[-1] + 1, worksheet.max_row + 1):
            values = [
                worksheet.cell(row=row_index, column=index).value
                for index in active_columns
            ]
            if not any(_has_value(value) for value in values):
                continue
            if rows and _looks_like_second_header(values):
                warnings.append(
                    "检测到可能的第二个独立表格，本轮仅分析首个表格。"
                )
                break
            rows.append(values)
            for position, column_index in enumerate(active_columns):
                cell = worksheet.cell(row=row_index, column=column_index)
                if _has_value(cell.value) and cell.number_format:
                    number_formats[headers[column_index - 1]].append(
                        cell.number_format
                    )

        data_frame = pd.DataFrame(
            rows,
            columns=[headers[index - 1] for index in active_columns],
        )
        non_empty_columns = [
            column
            for column in data_frame.columns
            if data_frame[column].map(_has_value).any()
        ]
        data_frame = data_frame.loc[:, non_empty_columns].copy()
        filters, report_period = _extract_filters(
            worksheet,
            header_rows[0],
            max_column,
        )

        return _ParsedAggregateWorkbook(
            sheet_name=extracted_schema.sheet_name,
            header_rows=header_rows,
            data_frame=data_frame,
            number_formats={
                column: tuple(number_formats.get(column, []))
                for column in data_frame.columns
            },
            report_period=report_period,
            filters=filters,
            warnings=tuple(warnings),
        )
    finally:
        workbook.close()


def _resolve_header_rows(
    worksheet: Any,
    header_row: int,
    max_column: int,
) -> tuple[int, ...]:
    if header_row <= 1:
        return (header_row,)
    previous_values = _expanded_row_values(
        worksheet,
        header_row - 1,
        max_column,
    )
    current_values = _expanded_row_values(
        worksheet,
        header_row,
        max_column,
    )
    previous_text = " ".join(
        str(value) for value in previous_values if _has_value(value)
    )
    current_semantic_count = sum(
        _looks_semantic_header(value)
        for value in current_values
        if _has_value(value)
    )
    has_horizontal_merge = any(
        merged.min_row == header_row - 1
        and merged.max_row == header_row - 1
        and merged.max_col > merged.min_col
        for merged in worksheet.merged_cells.ranges
    )
    has_group_hint = any(hint in previous_text for hint in GROUP_HEADER_HINTS)
    if current_semantic_count >= 2 and (has_horizontal_merge or has_group_hint):
        return (header_row - 1, header_row)
    return (header_row,)


def _expanded_row_values(
    worksheet: Any,
    row_index: int,
    max_column: int,
) -> list[Any]:
    values = [
        worksheet.cell(row=row_index, column=column_index).value
        for column_index in range(1, max_column + 1)
    ]
    for merged in worksheet.merged_cells.ranges:
        if not (merged.min_row <= row_index <= merged.max_row):
            continue
        merged_value = worksheet.cell(
            row=merged.min_row,
            column=merged.min_col,
        ).value
        for column_index in range(merged.min_col, merged.max_col + 1):
            if column_index <= max_column:
                values[column_index - 1] = merged_value
    return values


def _flatten_headers(header_rows: list[list[Any]]) -> list[str]:
    max_length = max((len(row) for row in header_rows), default=0)
    flattened: list[str] = []
    seen: dict[str, int] = {}
    for index in range(max_length):
        parts: list[str] = []
        for row in header_rows:
            value = row[index] if index < len(row) else None
            if not _has_value(value):
                continue
            text = str(value).strip()
            if text and text not in parts:
                parts.append(text)
        header = "_".join(parts)
        if not header:
            flattened.append("")
            continue
        seen[header] = seen.get(header, 0) + 1
        flattened.append(
            header if seen[header] == 1 else f"{header}__{seen[header]}"
        )
    return flattened


def _extract_filters(
    worksheet: Any,
    first_header_row: int,
    max_column: int,
) -> tuple[dict[str, str], str | None]:
    filters: dict[str, str] = {}
    report_period: str | None = None
    for row_index in range(1, first_header_row):
        values = [
            str(value).strip()
            for value in _expanded_row_values(
                worksheet,
                row_index,
                max_column,
            )
            if _has_value(value)
        ]
        if not values:
            continue
        if len(values) == 1 and ("：" in values[0] or ":" in values[0]):
            parts = re.split(r"[：:]", values[0], maxsplit=1)
            key, value = parts[0].strip(), parts[1].strip()
        elif len(values) >= 2:
            key, value = values[0], values[1]
        else:
            continue
        if not key or not value:
            continue
        if any(hint in key for hint in FILTER_HINTS):
            filters[key] = value
            if any(
                hint in key
                for hint in ("统计周期", "报表周期", "数据周期", "时间范围")
            ):
                report_period = value
    return filters, report_period


def _build_column_profiles(
    data_frame: pd.DataFrame,
    number_formats: dict[str, tuple[str, ...]],
) -> list[_ColumnProfile]:
    profiles: list[_ColumnProfile] = []
    for column_index, column in enumerate(data_frame.columns):
        values = [value for value in data_frame[column] if _has_value(value)]
        numeric_values = [
            numeric
            for value in values
            if (numeric := _raw_numeric(value)) is not None
        ]
        numeric_ratio = len(numeric_values) / len(values) if values else 0.0
        unique_ratio = (
            len({str(value).strip() for value in values}) / len(values)
            if values
            else 0.0
        )
        inferred_type = "numeric" if numeric_ratio >= 0.7 else "text"
        percentage_format = any(
            "%" in number_format
            for number_format in number_formats.get(str(column), ())
        ) or any(
            isinstance(value, str) and "%" in value for value in values
        )
        profiles.append(
            _ColumnProfile(
                source_column=str(column),
                column_index=column_index,
                inferred_type=inferred_type,
                numeric_ratio=numeric_ratio,
                unique_ratio=unique_ratio,
                percentage_format=percentage_format,
                value_range=_value_range_bucket(numeric_values),
            )
        )
    return profiles


def _recognize_semantic_fields(
    profiles: list[_ColumnProfile],
    *,
    fallback_resolver: AggregateFallbackResolver | None,
) -> tuple[list[_SemanticField], list[str]]:
    semantic_fields: list[_SemanticField] = []
    unresolved_profiles: list[_ColumnProfile] = []

    for profile in profiles:
        comparison = _recognize_comparison(profile)
        if comparison is not None:
            semantic_fields.append(comparison)
            continue
        metric = _recognize_metric(profile)
        if metric is not None:
            semantic_fields.append(metric)
            continue
        dimension = _recognize_dimension(profile)
        if dimension is not None:
            semantic_fields.append(dimension)
            continue
        unresolved_profiles.append(profile)

    if unresolved_profiles:
        resolver = fallback_resolver or _default_fallback_resolver
        try:
            ai_fields = resolver(
                [profile.ai_summary() for profile in unresolved_profiles]
            )
            semantic_fields.extend(
                _validated_ai_fields(ai_fields, unresolved_profiles)
            )
        except Exception as exc:
            logger.warning(
                "聚合字段 AI 兜底失败，保留规则识别结果：%r",
                exc,
            )

    recognized_sources = {field.source_column for field in semantic_fields}
    unrecognized_columns = [
        profile.source_column
        for profile in profiles
        if profile.source_column not in recognized_sources
    ]
    return semantic_fields, unrecognized_columns


def _default_fallback_resolver(
    profiles: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    from app.services.schema_mapper import map_aggregate_columns

    return map_aggregate_columns(profiles)


def _recognize_metric(profile: _ColumnProfile) -> _SemanticField | None:
    normalized = _normalize(profile.source_column)
    for rule in METRIC_RULES:
        if any(_normalize(alias) in normalized for alias in rule.aliases):
            return _SemanticField(
                source_column=profile.source_column,
                label=rule.label,
                semantic_key=rule.metric_key,
                role=rule.role,
                unit=rule.unit,
                aggregation=rule.aggregation,
                confidence="high",
                stage_order=rule.stage_order,
                semantic_role=rule.semantic_role,
            )
    if profile.percentage_format and profile.inferred_type == "numeric":
        return _SemanticField(
            source_column=profile.source_column,
            label=profile.source_column,
            semantic_key=_safe_semantic_key(profile.source_column),
            role="rate_metric",
            unit="ratio",
            aggregation="non_additive",
            confidence="medium",
        )
    return None


def _recognize_dimension(profile: _ColumnProfile) -> _SemanticField | None:
    normalized = _normalize(profile.source_column)
    for semantic_key, label, aliases in DIMENSION_RULES:
        if any(_normalize(alias) in normalized for alias in aliases):
            return _SemanticField(
                source_column=profile.source_column,
                label=label,
                semantic_key=semantic_key,
                role="dimension",
                unit="absolute",
                aggregation="non_additive",
                confidence="high",
            )
    if (
        profile.inferred_type == "text"
        and profile.column_index <= 2
        and 0 < profile.unique_ratio <= 1
        and not any(
            hint in normalized
            for hint in ("备注", "说明", "描述", "口径")
        )
    ):
        return _SemanticField(
            source_column=profile.source_column,
            label=profile.source_column,
            semantic_key=_safe_semantic_key(profile.source_column),
            role="dimension",
            unit="absolute",
            aggregation="non_additive",
            confidence="medium",
        )
    return None


def _recognize_comparison(
    profile: _ColumnProfile,
) -> _SemanticField | None:
    normalized = _normalize(profile.source_column)
    if not any(_normalize(hint) in normalized for hint in COMPARISON_HINTS):
        return None
    period = (
        "yoy"
        if "同比" in normalized or "同期" in normalized
        else "mom"
        if "环比" in normalized or "环期" in normalized or "上期" in normalized
        else None
    )
    target_rule = next(
        (
            rule
            for rule in METRIC_RULES
            if any(_normalize(alias) in normalized for alias in rule.aliases)
        ),
        None,
    )
    value_kind = (
        "delta"
        if any(hint in normalized for hint in COMPARISON_DELTA_HINTS)
        or "百分点" in normalized
        else "baseline"
        if "同期" in normalized or "上期" in normalized
        else "delta"
    )
    if value_kind == "baseline":
        unit = "absolute_value"
        comparison_type = period or "absolute_change"
    elif "百分点" in normalized or (
        "偏差" in normalized
        and target_rule is not None
        and target_rule.role == "rate_metric"
    ):
        unit = "percentage_point"
        comparison_type = "percentage_point_change"
    elif profile.percentage_format or any(
        hint in normalized for hint in ("率", "比例", "增幅")
    ):
        unit = "ratio_change"
        comparison_type = period or "absolute_change"
    else:
        unit = "absolute_change"
        comparison_type = "absolute_change"
    target_metric_key = target_rule.metric_key if target_rule else None
    semantic_key = "_".join(
        part
        for part in (
            target_metric_key or "metric",
            period or comparison_type,
            "baseline" if value_kind == "baseline" else "change",
        )
        if part
    )
    return _SemanticField(
        source_column=profile.source_column,
        label=profile.source_column,
        semantic_key=semantic_key,
        role="comparison_metric",
        unit=(
            "percentage_point"
            if unit == "percentage_point"
            else "ratio"
            if unit == "ratio_change"
            else "absolute"
        ),
        aggregation="non_additive",
        confidence="high" if target_metric_key else "medium",
        comparison_type=comparison_type,
        comparison_period=period,
        comparison_unit=unit,
        comparison_value_kind=value_kind,
        target_metric_key=target_metric_key,
    )


def _validated_ai_fields(
    payloads: list[dict[str, Any]],
    profiles: list[_ColumnProfile],
) -> list[_SemanticField]:
    profile_map = {profile.source_column: profile for profile in profiles}
    allowed_roles = {
        "dimension",
        "count_metric",
        "rate_metric",
        "amount_metric",
        "comparison_metric",
    }
    allowed_metric_keys = {
        rule.metric_key for rule in METRIC_RULES
    } | {key for key, _, _ in DIMENSION_RULES}
    confidence_map = {"high": "high", "medium": "medium", "low": "low"}
    fields: list[_SemanticField] = []
    used_sources: set[str] = set()
    for payload in payloads:
        source_column = payload.get("source_column")
        role = payload.get("role")
        semantic_key = payload.get("semantic_key")
        confidence = confidence_map.get(payload.get("confidence"))
        if (
            source_column not in profile_map
            or source_column in used_sources
            or role not in allowed_roles
            or semantic_key not in allowed_metric_keys
            or confidence is None
            or confidence == "low"
        ):
            continue
        rule = next(
            (
                candidate
                for candidate in METRIC_RULES
                if candidate.metric_key == semantic_key
            ),
            None,
        )
        if role == "dimension":
            label = next(
                (
                    dimension_label
                    for key, dimension_label, _ in DIMENSION_RULES
                    if key == semantic_key
                ),
                source_column,
            )
            fields.append(
                _SemanticField(
                    source_column=source_column,
                    label=label,
                    semantic_key=semantic_key,
                    role="dimension",
                    unit="absolute",
                    aggregation="non_additive",
                    confidence=confidence,
                )
            )
        elif rule is not None and role == rule.role:
            fields.append(
                _SemanticField(
                    source_column=source_column,
                    label=rule.label,
                    semantic_key=rule.metric_key,
                    role=rule.role,
                    unit=rule.unit,
                    aggregation=rule.aggregation,
                    confidence=confidence,
                    stage_order=rule.stage_order,
                    semantic_role=rule.semantic_role,
                )
            )
        used_sources.add(source_column)
    return fields


def _split_total_and_detail_rows(
    data_frame: pd.DataFrame,
    primary_dimension: _SemanticField | None,
) -> tuple[list[pd.Series], list[pd.Series]]:
    if primary_dimension is None:
        return [], [row for _, row in data_frame.iterrows()]
    total_rows: list[pd.Series] = []
    detail_rows: list[pd.Series] = []
    for _, row in data_frame.iterrows():
        dimension_value = row.get(primary_dimension.source_column)
        if not _has_value(dimension_value):
            continue
        normalized = _normalize(str(dimension_value))
        if normalized in TOTAL_ROW_LABELS:
            total_rows.append(row)
        else:
            detail_rows.append(row)
    return total_rows, detail_rows


def _select_scope_row(
    total_rows: list[pd.Series],
    detail_rows: list[pd.Series],
) -> tuple[pd.Series | None, str | None]:
    if total_rows:
        return total_rows[0], "total_row"
    if len(detail_rows) == 1:
        return detail_rows[0], "single_row"
    return None, None


def _deduplicate_metrics(
    metrics: list[_SemanticField],
) -> dict[str, _SemanticField]:
    lookup: dict[str, _SemanticField] = {}
    confidence_order = {"high": 3, "medium": 2, "low": 1}
    for metric in metrics:
        current = lookup.get(metric.semantic_key)
        if current is None or confidence_order[metric.confidence] > confidence_order[
            current.confidence
        ]:
            lookup[metric.semantic_key] = metric
    return lookup


def _build_funnel_definitions(
    metrics: list[_SemanticField],
) -> list[_SemanticField]:
    """Select exactly one canonical metric for each business funnel stage."""
    selected_by_order: dict[int, _SemanticField] = {}
    used_semantic_keys: set[str] = set()
    candidates = [
        metric
        for metric in metrics
        if metric.semantic_role == "funnel_stage"
        and metric.stage_order is not None
        and metric.confidence != "low"
    ]
    candidates.sort(key=_funnel_candidate_score, reverse=True)
    for metric in candidates:
        stage_order = metric.stage_order
        if (
            stage_order is None
            or stage_order in selected_by_order
            or metric.semantic_key in used_semantic_keys
        ):
            continue
        selected_by_order[stage_order] = metric
        used_semantic_keys.add(metric.semantic_key)
    return [selected_by_order[key] for key in sorted(selected_by_order)]


def _funnel_candidate_score(field: _SemanticField) -> tuple[int, int, int]:
    confidence_score = {"high": 3, "medium": 2, "low": 1}[field.confidence]
    rule = _metric_rule(field.semantic_key)
    normalized_source = _normalize(field.source_column)
    alias_score = 0
    for index, alias in enumerate(rule.aliases):
        if normalized_source == _normalize(alias):
            alias_score = len(rule.aliases) - index
            break
    main_payment_score = 1 if field.semantic_key == "payment_users" else 0
    return confidence_score, alias_score, main_payment_score


def _group_comparisons(
    comparisons: list[_SemanticField],
) -> dict[tuple[str | None, str | None], list[_SemanticField]]:
    grouped: dict[tuple[str | None, str | None], list[_SemanticField]] = {}
    for comparison in comparisons:
        key = (comparison.target_metric_key, comparison.comparison_period)
        grouped.setdefault(key, []).append(comparison)
    return grouped


def _build_kpis(
    scope_row: pd.Series | None,
    scope_source: str | None,
    detail_rows: list[pd.Series],
    metric_lookup: dict[str, _SemanticField],
    number_formats: dict[str, tuple[str, ...]],
) -> list[dict[str, Any]]:
    kpis: list[dict[str, Any]] = []
    values: dict[str, int | float] = {}
    sources: dict[str, str] = {}

    for metric_key in KPI_ORDER:
        metric = metric_lookup.get(metric_key)
        if metric is None:
            continue
        value: int | float | None = None
        source: str | None = None
        if scope_row is not None:
            value = _metric_value(scope_row, metric, number_formats)
            source = scope_source
        elif metric.aggregation == "sum" and detail_rows:
            row_values = [
                _metric_value(row, metric, number_formats)
                for row in detail_rows
            ]
            usable_values = [item for item in row_values if item is not None]
            if usable_values:
                value = round(float(sum(usable_values)), 2)
                source = "safe_sum"
        if value is not None and source is not None:
            values[metric_key] = value
            sources[metric_key] = source

    traffic = values.get("traffic_users")
    payment = values.get("payment_users")
    if isinstance(traffic, (int, float)) and isinstance(payment, (int, float)):
        if traffic > 0:
            values["payment_conversion_rate"] = round(payment / traffic, 6)
            sources["payment_conversion_rate"] = "derived"
    gmv = values.get("gmv")
    if (
        "average_order_value" not in values
        and isinstance(gmv, (int, float))
        and isinstance(payment, (int, float))
        and payment > 0
    ):
        values["average_order_value"] = round(gmv / payment, 2)
        sources["average_order_value"] = "derived"

    for metric_key in KPI_ORDER:
        if metric_key not in values:
            continue
        metric = metric_lookup.get(metric_key)
        if metric is None:
            if metric_key == "payment_conversion_rate":
                metric = _rule_to_field(_metric_rule(metric_key), "derived")
            elif metric_key == "average_order_value":
                metric = _rule_to_field(_metric_rule(metric_key), "derived")
            else:
                continue
        kpis.append(
            {
                "metric_key": metric_key,
                "label": metric.label,
                "value": values[metric_key],
                "unit": metric.unit,
                "aggregation": metric.aggregation,
                "source": sources[metric_key],
            }
        )
    return kpis


def _build_dimension_performance(
    detail_rows: list[pd.Series],
    primary_dimension: _SemanticField | None,
    metric_lookup: dict[str, _SemanticField],
    comparison_lookup: dict[
        tuple[str | None, str | None],
        list[_SemanticField],
    ],
    number_formats: dict[str, tuple[str, ...]],
) -> list[dict[str, Any]]:
    if primary_dimension is None:
        return []
    performance: list[dict[str, Any]] = []
    for row in detail_rows:
        dimension_value = row.get(primary_dimension.source_column)
        if not _has_value(dimension_value):
            continue
        traffic = _lookup_metric_value(
            row, "traffic_users", metric_lookup, number_formats
        )
        appointment = _lookup_metric_value(
            row, "appointment_users", metric_lookup, number_formats
        )
        payment = _lookup_metric_value(
            row, "payment_users", metric_lookup, number_formats
        )
        gmv = _lookup_metric_value(row, "gmv", metric_lookup, number_formats)
        average_order_value = _lookup_metric_value(
            row,
            "average_order_value",
            metric_lookup,
            number_formats,
        )
        if (
            average_order_value is None
            and gmv is not None
            and payment is not None
            and payment > 0
        ):
            average_order_value = round(gmv / payment, 2)
        if traffic is not None and payment is not None and traffic > 0:
            conversion_rate = round(payment / traffic, 6)
        else:
            conversion_rate = _lookup_metric_value(
                row,
                "payment_conversion_rate",
                metric_lookup,
                number_formats,
            )
        yoy, yoy_unit = _resolve_row_comparison(
            row,
            "yoy",
            metric_lookup,
            comparison_lookup,
            number_formats,
        )
        mom, mom_unit = _resolve_row_comparison(
            row,
            "mom",
            metric_lookup,
            comparison_lookup,
            number_formats,
        )
        performance.append(
            {
                "dimension_value": str(dimension_value).strip(),
                "traffic_users": _as_int(traffic),
                "appointment_users": _as_int(appointment),
                "payment_users": _as_int(payment),
                "conversion_rate": _as_float(conversion_rate),
                "gmv": _as_float(gmv),
                "average_order_value": _as_float(average_order_value),
                "yoy": yoy,
                "mom": mom,
                "yoy_unit": yoy_unit,
                "mom_unit": mom_unit,
                "supplemental_outcomes": _build_supplemental_outcomes(
                    row,
                    metric_lookup,
                    number_formats,
                ),
            }
        )
    return performance


def _build_supplemental_outcomes(
    row: pd.Series,
    metric_lookup: dict[str, _SemanticField],
    number_formats: dict[str, tuple[str, ...]],
) -> list[dict[str, Any]]:
    outcomes: list[dict[str, Any]] = []
    for metric in metric_lookup.values():
        if (
            metric.semantic_role != "outcome_metric"
            or metric.role != "count_metric"
        ):
            continue
        value = _metric_value(row, metric, number_formats)
        if value is None:
            continue
        outcomes.append(
            {
                "metric_key": metric.semantic_key,
                "label": metric.label,
                "value": value,
                "unit": metric.unit,
            }
        )
    return outcomes


def _build_dynamic_funnel(
    scope_row: pd.Series | None,
    scope_source: str | None,
    funnel_definitions: list[_SemanticField],
    comparison_lookup: dict[
        tuple[str | None, str | None],
        list[_SemanticField],
    ],
    number_formats: dict[str, tuple[str, ...]],
    primary_dimension: _SemanticField | None,
) -> tuple[dict[str, Any], list[str]]:
    warnings: list[str] = []
    if scope_row is None:
        return {"scope_dimension_value": None, "stages": []}, warnings
    if scope_source == "total_row" and primary_dimension is not None:
        scope_value = str(scope_row.get(primary_dimension.source_column)).strip()
    elif scope_source == "single_row" and primary_dimension is not None:
        scope_value = str(scope_row.get(primary_dimension.source_column)).strip()
    else:
        scope_value = None
    stages: list[dict[str, Any]] = []
    previous_count: int | None = None
    previous_semantic_key: str | None = None
    for definition in funnel_definitions:
        if definition.semantic_key == previous_semantic_key:
            continue
        value = _metric_value(scope_row, definition, number_formats)
        current_count = _as_int(value)
        if current_count is None:
            continue
        if previous_count is None:
            conversion_rate = None
            dropoff_count = None
        else:
            conversion_rate = (
                round(current_count / previous_count, 6)
                if previous_count > 0
                else None
            )
            dropoff_count = previous_count - current_count
            if current_count > previous_count:
                warnings.append(
                    f"漏斗阶段“{definition.label}”人数高于前序阶段，"
                    "请确认指标口径是否一致。"
                )
        yoy, yoy_unit = _resolve_comparison_for_target(
            scope_row,
            definition.semantic_key,
            "yoy",
            definition,
            comparison_lookup,
            number_formats,
        )
        mom, mom_unit = _resolve_comparison_for_target(
            scope_row,
            definition.semantic_key,
            "mom",
            definition,
            comparison_lookup,
            number_formats,
        )
        stages.append(
            {
                "metric_key": definition.semantic_key,
                "label": definition.label,
                "user_count": current_count,
                "conversion_rate_from_previous": conversion_rate,
                "dropoff_count": dropoff_count,
                "yoy": yoy,
                "mom": mom,
                "yoy_unit": yoy_unit,
                "mom_unit": mom_unit,
            }
        )
        previous_count = current_count
        previous_semantic_key = definition.semantic_key
    return {"scope_dimension_value": scope_value, "stages": stages}, warnings


def _build_dimension_funnel_diagnostics(
    detail_rows: list[pd.Series],
    primary_dimension: _SemanticField | None,
    funnel_definitions: list[_SemanticField],
    metric_lookup: dict[str, _SemanticField],
    comparison_lookup: dict[
        tuple[str | None, str | None],
        list[_SemanticField],
    ],
    number_formats: dict[str, tuple[str, ...]],
    dimension_performance: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    if primary_dimension is None:
        return []

    summaries: list[dict[str, Any]] = []
    for row in detail_rows:
        dimension_value = row.get(primary_dimension.source_column)
        if not _has_value(dimension_value):
            continue
        stages = _build_row_funnel_stages(
            row,
            funnel_definitions,
            metric_lookup,
            comparison_lookup,
            number_formats,
        )
        final_rate = _derive_final_conversion_rate(
            row,
            metric_lookup,
            number_formats,
        )
        final_yoy, final_yoy_unit = _resolve_rate_comparison(
            row,
            current_rate=final_rate,
            rate_metric_key="payment_conversion_rate",
            from_metric_key="traffic_users",
            to_metric_key="payment_users",
            period="yoy",
            metric_lookup=metric_lookup,
            comparison_lookup=comparison_lookup,
            number_formats=number_formats,
        )
        final_mom, final_mom_unit = _resolve_rate_comparison(
            row,
            current_rate=final_rate,
            rate_metric_key="payment_conversion_rate",
            from_metric_key="traffic_users",
            to_metric_key="payment_users",
            period="mom",
            metric_lookup=metric_lookup,
            comparison_lookup=comparison_lookup,
            number_formats=number_formats,
        )
        summaries.append(
            {
                "dimension_value": str(dimension_value).strip(),
                "final_conversion_rate": final_rate,
                "final_conversion_yoy": final_yoy,
                "final_conversion_mom": final_mom,
                "final_conversion_yoy_unit": final_yoy_unit,
                "final_conversion_mom_unit": final_mom_unit,
                "stages": stages,
                "best_improving_stage": _select_stage_movement(
                    stages,
                    direction="positive",
                ),
                "largest_declining_stage": _select_stage_movement(
                    stages,
                    direction="negative",
                ),
                "weakest_stage": None,
                "diagnosis_level": "stable",
            }
        )

    peer_rates: dict[tuple[str, str], list[float]] = {}
    performance_lookup = {
        item["dimension_value"]: item for item in dimension_performance
    }
    traffic_median = _median_optional(
        item["traffic_users"] for item in dimension_performance
    )
    payment_median = _median_optional(
        item["payment_users"] for item in dimension_performance
    )
    for summary in summaries:
        for stage in summary["stages"]:
            current_rate = stage["current_conversion_rate"]
            if current_rate is None:
                continue
            key = (stage["from_metric_key"], stage["to_metric_key"])
            peer_rates.setdefault(key, []).append(current_rate)

    for summary in summaries:
        summary["weakest_stage"] = _select_weakest_stage(
            summary["stages"],
            peer_rates,
        )
        summary["diagnosis_level"] = _resolve_dimension_diagnosis_level(
            summary,
            performance_lookup.get(summary["dimension_value"], {}),
            traffic_median=traffic_median,
            payment_median=payment_median,
        )
    return summaries


def _build_row_funnel_stages(
    row: pd.Series,
    funnel_definitions: list[_SemanticField],
    metric_lookup: dict[str, _SemanticField],
    comparison_lookup: dict[
        tuple[str | None, str | None],
        list[_SemanticField],
    ],
    number_formats: dict[str, tuple[str, ...]],
) -> list[dict[str, Any]]:
    available_stages: list[tuple[_SemanticField, int]] = []
    used_semantic_keys: set[str] = set()
    for definition in funnel_definitions:
        if definition.semantic_key in used_semantic_keys:
            continue
        count = _as_int(_metric_value(row, definition, number_formats))
        if count is not None:
            available_stages.append((definition, count))
            used_semantic_keys.add(definition.semantic_key)

    stages: list[dict[str, Any]] = []
    for (from_stage, from_count), (to_stage, to_count) in zip(
        available_stages,
        available_stages[1:],
    ):
        if (
            from_stage.semantic_key == to_stage.semantic_key
            or from_stage.stage_order == to_stage.stage_order
        ):
            continue
        rate_metric_key = FUNNEL_RATE_KEYS.get(
            (from_stage.semantic_key, to_stage.semantic_key)
        )
        current_rate = (
            round(to_count / from_count, 6)
            if from_count > 0
            else None
        )
        if current_rate is None and rate_metric_key is not None:
            current_rate = _as_float(
                _lookup_metric_value(
                    row,
                    rate_metric_key,
                    metric_lookup,
                    number_formats,
                )
            )
        yoy, yoy_unit = _resolve_rate_comparison(
            row,
            current_rate=current_rate,
            rate_metric_key=rate_metric_key,
            from_metric_key=from_stage.semantic_key,
            to_metric_key=to_stage.semantic_key,
            period="yoy",
            metric_lookup=metric_lookup,
            comparison_lookup=comparison_lookup,
            number_formats=number_formats,
        )
        mom, mom_unit = _resolve_rate_comparison(
            row,
            current_rate=current_rate,
            rate_metric_key=rate_metric_key,
            from_metric_key=from_stage.semantic_key,
            to_metric_key=to_stage.semantic_key,
            period="mom",
            metric_lookup=metric_lookup,
            comparison_lookup=comparison_lookup,
            number_formats=number_formats,
        )
        stages.append(
            {
                "from_metric_key": from_stage.semantic_key,
                "from_label": from_stage.label,
                "to_metric_key": to_stage.semantic_key,
                "to_label": to_stage.label,
                "current_conversion_rate": current_rate,
                "yoy_delta": yoy,
                "mom_delta": mom,
                "yoy_unit": yoy_unit,
                "mom_unit": mom_unit,
            }
        )
    return stages


def _derive_final_conversion_rate(
    row: pd.Series,
    metric_lookup: dict[str, _SemanticField],
    number_formats: dict[str, tuple[str, ...]],
) -> float | None:
    traffic = _lookup_metric_value(
        row,
        "traffic_users",
        metric_lookup,
        number_formats,
    )
    payment = _lookup_metric_value(
        row,
        "payment_users",
        metric_lookup,
        number_formats,
    )
    if traffic is not None and payment is not None and traffic > 0:
        return round(payment / traffic, 6)
    return _as_float(
        _lookup_metric_value(
            row,
            "payment_conversion_rate",
            metric_lookup,
            number_formats,
        )
    )


def _resolve_rate_comparison(
    row: pd.Series,
    *,
    current_rate: float | None,
    rate_metric_key: str | None,
    from_metric_key: str,
    to_metric_key: str,
    period: str,
    metric_lookup: dict[str, _SemanticField],
    comparison_lookup: dict[
        tuple[str | None, str | None],
        list[_SemanticField],
    ],
    number_formats: dict[str, tuple[str, ...]],
) -> tuple[float | None, str | None]:
    if rate_metric_key is not None:
        comparisons = comparison_lookup.get((rate_metric_key, period), [])
        if comparisons:
            comparison = comparisons[0]
            if comparison.comparison_value_kind == "baseline":
                baseline_rate = _coerce_numeric(
                    row.get(comparison.source_column),
                    "ratio",
                    number_formats.get(comparison.source_column, ()),
                )
                if current_rate is not None and baseline_rate is not None:
                    return (
                        round(current_rate - baseline_rate, 6),
                        "percentage_point",
                    )
            else:
                value = _comparison_value(
                    row,
                    comparison,
                    current_metric=metric_lookup.get(rate_metric_key),
                    number_formats=number_formats,
                )
                if value is not None and comparison.comparison_unit in {
                    "ratio_change",
                    "percentage_point",
                }:
                    return value, comparison.comparison_unit

    from_baseline = _comparison_baseline_value(
        row,
        from_metric_key,
        period,
        comparison_lookup,
        number_formats,
    )
    to_baseline = _comparison_baseline_value(
        row,
        to_metric_key,
        period,
        comparison_lookup,
        number_formats,
    )
    if (
        current_rate is not None
        and from_baseline is not None
        and from_baseline > 0
        and to_baseline is not None
    ):
        baseline_rate = to_baseline / from_baseline
        return round(current_rate - baseline_rate, 6), "percentage_point"
    return None, None


def _comparison_baseline_value(
    row: pd.Series,
    metric_key: str,
    period: str,
    comparison_lookup: dict[
        tuple[str | None, str | None],
        list[_SemanticField],
    ],
    number_formats: dict[str, tuple[str, ...]],
) -> float | None:
    comparisons = comparison_lookup.get((metric_key, period), [])
    baseline = next(
        (
            comparison
            for comparison in comparisons
            if comparison.comparison_value_kind == "baseline"
        ),
        None,
    )
    if baseline is None:
        return None
    return _coerce_numeric(
        row.get(baseline.source_column),
        "absolute",
        number_formats.get(baseline.source_column, ()),
    )


def _select_stage_movement(
    stages: list[dict[str, Any]],
    *,
    direction: str,
) -> dict[str, Any] | None:
    candidates: list[dict[str, Any]] = []
    for stage in stages:
        for comparison, delta_key, unit_key in (
            ("yoy", "yoy_delta", "yoy_unit"),
            ("mom", "mom_delta", "mom_unit"),
        ):
            delta = stage[delta_key]
            unit = stage[unit_key]
            if delta is None or unit is None:
                continue
            if direction == "positive" and delta < 0.02:
                continue
            if direction == "negative" and delta > -0.02:
                continue
            candidates.append(
                {
                    "from_metric_key": stage["from_metric_key"],
                    "from_label": stage["from_label"],
                    "to_metric_key": stage["to_metric_key"],
                    "to_label": stage["to_label"],
                    "delta": delta,
                    "comparison": comparison,
                    "unit": unit,
                }
            )
    if not candidates:
        return None
    key = (lambda item: item["delta"])
    return (
        max(candidates, key=key)
        if direction == "positive"
        else min(candidates, key=key)
    )


def _select_weakest_stage(
    stages: list[dict[str, Any]],
    peer_rates: dict[tuple[str, str], list[float]],
) -> dict[str, Any] | None:
    row_rates = [
        stage["current_conversion_rate"]
        for stage in stages
        if stage["current_conversion_rate"] is not None
    ]
    row_median = median(row_rates) if row_rates else None
    candidates: list[tuple[float, dict[str, Any], float | None]] = []
    for stage in stages:
        current_rate = stage["current_conversion_rate"]
        if current_rate is None:
            continue
        key = (stage["from_metric_key"], stage["to_metric_key"])
        peers = peer_rates.get(key, [])
        peer_median = median(peers) if len(peers) >= 2 else None
        peer_gap = (
            max(peer_median - current_rate, 0)
            if peer_median is not None
            else 0
        )
        row_gap = (
            max(row_median - current_rate, 0)
            if row_median is not None
            else 0
        )
        negative_change = max(
            -float(stage["yoy_delta"] or 0),
            -float(stage["mom_delta"] or 0),
            0,
        )
        score = peer_gap + negative_change + row_gap * 0.35
        if peer_gap < 0.03 and negative_change < 0.03 and row_gap < 0.08:
            continue
        candidates.append((score, stage, peer_median))
    if not candidates:
        return None
    _, stage, peer_median = max(candidates, key=lambda item: item[0])
    return {
        "from_metric_key": stage["from_metric_key"],
        "from_label": stage["from_label"],
        "to_metric_key": stage["to_metric_key"],
        "to_label": stage["to_label"],
        "current_conversion_rate": stage["current_conversion_rate"],
        "peer_median_conversion_rate": peer_median,
        "yoy_delta": stage["yoy_delta"],
        "mom_delta": stage["mom_delta"],
    }


def _resolve_dimension_diagnosis_level(
    summary: dict[str, Any],
    performance: dict[str, Any],
    *,
    traffic_median: float | None,
    payment_median: float | None,
) -> str:
    final_changes = [
        value
        for value in (
            summary["final_conversion_yoy"],
            summary["final_conversion_mom"],
        )
        if value is not None
    ]
    stage_decline = (
        summary["largest_declining_stage"]["delta"]
        if summary["largest_declining_stage"]
        else None
    )
    stage_improvement = (
        summary["best_improving_stage"]["delta"]
        if summary["best_improving_stage"]
        else None
    )
    negative_magnitude = abs(
        min([0.0, *final_changes, stage_decline or 0.0])
    )
    positive_magnitude = max(
        [0.0, *final_changes, stage_improvement or 0.0]
    )
    traffic = performance.get("traffic_users")
    payment = performance.get("payment_users")
    material_scale = bool(
        (
            traffic is not None
            and traffic_median is not None
            and traffic >= traffic_median
        )
        or (
            payment is not None
            and payment_median is not None
            and payment >= payment_median
        )
    )

    if negative_magnitude >= 0.1 or (
        negative_magnitude >= 0.06 and material_scale
    ):
        return "high_priority"
    if negative_magnitude >= 0.03 or (
        summary["weakest_stage"] is not None
        and negative_magnitude >= 0.02
    ):
        return "attention"
    if positive_magnitude >= 0.05:
        return "improving"
    return "stable"


def _median_optional(values: Iterable[int | float | None]) -> float | None:
    usable_values = [float(value) for value in values if value is not None]
    return float(median(usable_values)) if usable_values else None


def _apply_final_comparisons_to_performance(
    performance: list[dict[str, Any]],
    summaries: list[dict[str, Any]],
) -> None:
    summary_lookup = {
        summary["dimension_value"]: summary for summary in summaries
    }
    for item in performance:
        summary = summary_lookup.get(item["dimension_value"])
        if summary is None:
            continue
        item["conversion_rate"] = summary["final_conversion_rate"]
        item["yoy"] = summary["final_conversion_yoy"]
        item["mom"] = summary["final_conversion_mom"]
        item["yoy_unit"] = summary["final_conversion_yoy_unit"]
        item["mom_unit"] = summary["final_conversion_mom_unit"]


def _build_diagnostics(
    performance: list[dict[str, Any]],
    dimension_funnels: list[dict[str, Any]],
    detail_rows: list[pd.Series],
    primary_dimension: _SemanticField | None,
    metric_lookup: dict[str, _SemanticField],
    comparison_lookup: dict[
        tuple[str | None, str | None],
        list[_SemanticField],
    ],
    number_formats: dict[str, tuple[str, ...]],
) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for summary in dimension_funnels:
        dimension_value = summary["dimension_value"]
        final_rate = summary["final_conversion_rate"]
        improvement = summary["best_improving_stage"]
        decline = summary["largest_declining_stage"]
        final_trend_added = False
        comparisons = [
            (
                "yoy",
                summary["final_conversion_yoy"],
                summary["final_conversion_yoy_unit"],
            ),
            (
                "mom",
                summary["final_conversion_mom"],
                summary["final_conversion_mom_unit"],
            ),
        ]
        usable_comparisons = [
            item for item in comparisons if item[1] is not None
        ]
        if usable_comparisons:
            period, change, unit = max(
                usable_comparisons,
                key=lambda item: abs(item[1]),
            )
            if abs(change) >= 0.02:
                period_label = "同比" if period == "yoy" else "环比"
                improving = change > 0
                positive_node_text = (
                    f"主要正向节点为{_business_stage_label(improvement['from_label'])}→"
                    f"{_business_stage_label(improvement['to_label'])}，"
                    f"{('同比' if improvement['comparison'] == 'yoy' else '环比')}"
                    f"{_comparison_text(improvement['delta'], improvement['unit'])}。"
                    if improving and improvement is not None
                    else ""
                )
                candidates.append(
                    {
                        "diagnostic_type": (
                            "conversion_trend_improvement"
                            if improving
                            else "yoy_decline"
                            if period == "yoy"
                            else "mom_decline"
                        ),
                        "title": (
                            f"{dimension_value}支付转化率{period_label}"
                            f"{'明显改善' if improving else '明显下滑'}"
                        ),
                        "evidence": (
                            f"当前支付转化率 "
                            f"{_optional_percent_text(final_rate)}，"
                            f"{period_label}"
                            f"{_comparison_text(change, unit)}。"
                            f"{positive_node_text}"
                        ),
                        "severity": _severity_for_change(change),
                        "dimension_value": dimension_value,
                        "metric_key": "payment_conversion_rate",
                        "_score": abs(change) * 100 + 100,
                        "_dedupe_key": "final_conversion_trend",
                    }
                )
                final_trend_added = True

        if decline is not None:
            period_label = "同比" if decline["comparison"] == "yoy" else "环比"
            candidates.append(
                {
                    "diagnostic_type": "stage_decline",
                    "title": (
                        f"{dimension_value}的"
                        f"{_business_stage_label(decline['from_label'])}→"
                        f"{_business_stage_label(decline['to_label'])}是主要拖累节点"
                    ),
                    "evidence": (
                        f"该环节转化{period_label}"
                        f"{_comparison_text(decline['delta'], decline['unit'])}。"
                    ),
                    "severity": _severity_for_change(decline["delta"]),
                    "dimension_value": dimension_value,
                    "metric_key": decline["to_metric_key"],
                    "_score": abs(decline["delta"]) * 100 + 7,
                    "_dedupe_key": (
                        f"stage:{decline['from_metric_key']}:"
                        f"{decline['to_metric_key']}"
                    ),
                }
            )

        if improvement is not None and not final_trend_added:
            period_label = (
                "同比" if improvement["comparison"] == "yoy" else "环比"
            )
            candidates.append(
                {
                    "diagnostic_type": "stage_improvement",
                    "title": (
                        f"{dimension_value}的"
                        f"{_business_stage_label(improvement['from_label'])}→"
                        f"{_business_stage_label(improvement['to_label'])}是主要正向节点"
                    ),
                    "evidence": (
                        f"该环节转化{period_label}"
                        f"{_comparison_text(improvement['delta'], improvement['unit'])}。"
                    ),
                    "severity": _severity_for_change(improvement["delta"]),
                    "dimension_value": dimension_value,
                    "metric_key": improvement["to_metric_key"],
                    "_score": abs(improvement["delta"]) * 100 + 4,
                    "_dedupe_key": (
                        f"stage:{improvement['from_metric_key']}:"
                        f"{improvement['to_metric_key']}"
                    ),
                }
            )

        weakest = summary["weakest_stage"]
        if weakest is not None:
            peer_text = (
                f"，同环节品类中位数为 "
                f"{_percent_text(weakest['peer_median_conversion_rate'])}"
                if weakest["peer_median_conversion_rate"] is not None
                else ""
            )
            candidates.append(
                {
                    "diagnostic_type": "weakest_stage",
                    "title": (
                        f"{dimension_value}的"
                        f"{_business_stage_label(weakest['from_label'])}→"
                        f"{_business_stage_label(weakest['to_label'])}是当前薄弱环节"
                    ),
                    "evidence": (
                        "当前阶段转化率为 "
                        f"{_percent_text(weakest['current_conversion_rate'])}"
                        f"{peer_text}。"
                    ),
                    "severity": "medium",
                    "dimension_value": dimension_value,
                    "metric_key": weakest["to_metric_key"],
                    "_score": 5,
                    "_dedupe_key": (
                        f"stage:{weakest['from_metric_key']}:"
                        f"{weakest['to_metric_key']}"
                    ),
                }
            )

    traffic_values = [
        item["traffic_users"]
        for item in performance
        if item["traffic_users"] is not None
    ]
    conversion_values = [
        item["conversion_rate"]
        for item in performance
        if item["conversion_rate"] is not None
    ]
    if len(traffic_values) >= 3 and len(conversion_values) >= 3:
        traffic_median = median(traffic_values)
        conversion_median = median(conversion_values)
        for item in performance:
            traffic = item["traffic_users"]
            conversion = item["conversion_rate"]
            if traffic is None or conversion is None:
                continue
            if (
                traffic >= traffic_median * 1.2
                and conversion <= conversion_median - 0.03
            ):
                candidates.append(
                    {
                        "diagnostic_type": "high_traffic_low_conversion",
                        "title": f"{item['dimension_value']}流量较高但转化偏低",
                        "evidence": (
                            f"浏览用户 {traffic}，高于同维度中位数 "
                            f"{traffic_median:,.0f}；支付转化率 "
                            f"{_percent_text(conversion)}，低于中位水平 "
                            f"{_percent_text(conversion_median)}。"
                        ),
                        "severity": (
                            "high"
                            if conversion <= conversion_median - 0.08
                            else "medium"
                        ),
                        "dimension_value": item["dimension_value"],
                        "metric_key": "payment_conversion_rate",
                        "_score": (conversion_median - conversion) * 100 + 5,
                        "_dedupe_key": "scale_conversion_position",
                    }
                )
            elif (
                traffic <= traffic_median * 0.7
                and conversion >= conversion_median + 0.03
            ):
                candidates.append(
                    {
                        "diagnostic_type": "high_conversion_low_traffic",
                        "title": f"{item['dimension_value']}具备审慎扩量机会",
                        "evidence": (
                            f"支付转化率 {_percent_text(conversion)}，高于同维度"
                            f"中位水平 {_percent_text(conversion_median)}；浏览用户 "
                            f"{traffic}，低于中位数 {traffic_median:,.0f}。"
                        ),
                        "severity": "medium",
                        "dimension_value": item["dimension_value"],
                        "metric_key": "traffic_users",
                        "_score": (conversion - conversion_median) * 100 + 3,
                        "_dedupe_key": "scale_conversion_position",
                    }
                )
    for diagnostic in _build_gmv_payment_mismatch_diagnostics(
        detail_rows,
        primary_dimension,
        metric_lookup,
        comparison_lookup,
        number_formats,
    ):
        candidates.append(
            {
                **diagnostic,
                "_score": 6,
                "_dedupe_key": "gmv_payment_mismatch",
            }
        )
    return _rank_and_limit_diagnostics(candidates)


def _build_gmv_payment_mismatch_diagnostics(
    rows: list[pd.Series],
    primary_dimension: _SemanticField | None,
    metric_lookup: dict[str, _SemanticField],
    comparison_lookup: dict[
        tuple[str | None, str | None],
        list[_SemanticField],
    ],
    number_formats: dict[str, tuple[str, ...]],
) -> list[dict[str, Any]]:
    if primary_dimension is None:
        return []
    results: list[dict[str, Any]] = []
    for row in rows:
        dimension_value = row.get(primary_dimension.source_column)
        if not _has_value(dimension_value):
            continue
        for period in ("yoy", "mom"):
            payment_change, payment_unit = _resolve_comparison_for_target(
                row,
                "payment_users",
                period,
                metric_lookup.get("payment_users"),
                comparison_lookup,
                number_formats,
            )
            gmv_change, gmv_unit = _resolve_comparison_for_target(
                row,
                "gmv",
                period,
                metric_lookup.get("gmv"),
                comparison_lookup,
                number_formats,
            )
            if (
                payment_change is None
                or gmv_change is None
                or payment_change == 0
                or gmv_change == 0
                or math.copysign(1, payment_change) == math.copysign(1, gmv_change)
            ):
                continue
            period_label = "同比" if period == "yoy" else "环比"
            results.append(
                {
                    "diagnostic_type": "gmv_payment_mismatch",
                    "title": f"{dimension_value}的 GMV 与支付用户变化方向不一致",
                    "evidence": (
                        f"{period_label}支付用户 "
                        f"{_comparison_text(payment_change, payment_unit)}，"
                        f"GMV {_comparison_text(gmv_change, gmv_unit)}。"
                    ),
                    "severity": "medium",
                    "dimension_value": str(dimension_value).strip(),
                    "metric_key": "gmv",
                }
            )
    return results


def _build_opportunities(
    performance: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    opportunities: list[dict[str, Any]] = []
    gmv_items = [item for item in performance if item["gmv"] is not None]
    if gmv_items:
        top_gmv = max(gmv_items, key=lambda item: item["gmv"])
        opportunities.append(
            {
                "opportunity_type": "high_gmv",
                "title": f"{top_gmv['dimension_value']}贡献较高 GMV",
                "evidence": f"当前 GMV 为 {top_gmv['gmv']:.2f}。",
                "dimension_value": top_gmv["dimension_value"],
                "metric_key": "gmv",
            }
        )
    return opportunities


def _build_business_insights(
    dimension_funnels: list[dict[str, Any]],
    diagnostics: list[dict[str, Any]],
    opportunities: list[dict[str, Any]],
    performance: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    diagnostics_by_dimension: dict[str, list[dict[str, Any]]] = {}
    for diagnostic in diagnostics:
        dimension_value = diagnostic.get("dimension_value")
        if dimension_value:
            diagnostics_by_dimension.setdefault(dimension_value, []).append(
                diagnostic
            )
    opportunities_by_dimension: dict[str, list[dict[str, Any]]] = {}
    for opportunity in opportunities:
        opportunities_by_dimension.setdefault(
            opportunity["dimension_value"],
            [],
        ).append(opportunity)
    performance_lookup = {
        item["dimension_value"]: item for item in performance
    }

    ranked_insights: list[tuple[float, dict[str, Any]]] = []
    for summary in dimension_funnels:
        dimension_value = summary["dimension_value"]
        dimension_diagnostics = diagnostics_by_dimension.get(
            dimension_value,
            [],
        )
        dimension_opportunities = opportunities_by_dimension.get(
            dimension_value,
            [],
        )
        if not _has_meaningful_business_signal(
            summary,
            dimension_diagnostics,
            dimension_opportunities,
        ):
            continue

        positive_signal = _build_positive_signal(
            summary,
            dimension_diagnostics,
            dimension_opportunities,
        )
        risk_signal = _build_risk_signal(
            summary,
            dimension_diagnostics,
        )
        priority = _resolve_business_insight_priority(summary)
        evidence = _build_business_evidence(summary)
        if not evidence:
            evidence = [
                item["evidence"]
                for item in dimension_diagnostics[:2]
                if item.get("evidence")
            ]
        core_judgement = _build_core_judgement(
            summary,
            positive_signal,
            risk_signal,
            dimension_diagnostics,
            dimension_opportunities,
        )
        traffic = performance_lookup.get(dimension_value, {}).get(
            "traffic_users"
        )
        rank_score = _business_insight_rank_score(
            summary,
            priority,
            traffic,
        )
        ranked_insights.append(
            (
                rank_score,
                {
                    "dimension_value": dimension_value,
                    "core_judgement": core_judgement,
                    "positive_signal": positive_signal,
                    "risk_signal": risk_signal,
                    "key_evidence": evidence[:2],
                    "priority": priority,
                },
            )
        )

    ranked_insights.sort(key=lambda item: item[0], reverse=True)
    return [insight for _, insight in ranked_insights[:5]]


def _has_meaningful_business_signal(
    summary: dict[str, Any],
    diagnostics: list[dict[str, Any]],
    opportunities: list[dict[str, Any]],
) -> bool:
    final_movements = (
        summary["final_conversion_yoy"],
        summary["final_conversion_mom"],
    )
    return bool(
        diagnostics
        or opportunities
        or summary["best_improving_stage"]
        or summary["largest_declining_stage"]
        or summary["weakest_stage"]
        or any(
            value is not None and abs(value) >= 0.02
            for value in final_movements
        )
    )


def _build_positive_signal(
    summary: dict[str, Any],
    diagnostics: list[dict[str, Any]],
    opportunities: list[dict[str, Any]],
) -> str | None:
    improvement = summary["best_improving_stage"]
    if improvement is not None:
        period_label = (
            "同比" if improvement["comparison"] == "yoy" else "环比"
        )
        return (
            f"{_business_stage_label(improvement['from_label'])}→"
            f"{_business_stage_label(improvement['to_label'])}"
            f"{period_label}{_comparison_text(improvement['delta'], improvement['unit'])}"
        )
    for period_label, value, unit in (
        (
            "同比",
            summary["final_conversion_yoy"],
            summary["final_conversion_yoy_unit"],
        ),
        (
            "环比",
            summary["final_conversion_mom"],
            summary["final_conversion_mom_unit"],
        ),
    ):
        if value is not None and value >= 0.02:
            return f"支付转化率{period_label}{_comparison_text(value, unit)}"
    expansion = next(
        (
            item
            for item in diagnostics
            if item["diagnostic_type"] == "high_conversion_low_traffic"
        ),
        None,
    )
    if expansion is not None:
        return expansion["title"]
    if opportunities:
        return opportunities[0]["title"]
    return None


def _build_risk_signal(
    summary: dict[str, Any],
    diagnostics: list[dict[str, Any]],
) -> str | None:
    decline = summary["largest_declining_stage"]
    if decline is not None:
        period_label = "同比" if decline["comparison"] == "yoy" else "环比"
        return (
            f"{_business_stage_label(decline['from_label'])}→"
            f"{_business_stage_label(decline['to_label'])}"
            f"{period_label}{_comparison_text(decline['delta'], decline['unit'])}"
        )
    weakest = summary["weakest_stage"]
    if weakest is not None:
        return (
            f"{_business_stage_label(weakest['from_label'])}→"
            f"{_business_stage_label(weakest['to_label'])}当前转化率"
            f"{_percent_text(weakest['current_conversion_rate'])}"
        )
    high_traffic_risk = next(
        (
            item
            for item in diagnostics
            if item["diagnostic_type"] == "high_traffic_low_conversion"
        ),
        None,
    )
    return high_traffic_risk["title"] if high_traffic_risk else None


def _build_core_judgement(
    summary: dict[str, Any],
    positive_signal: str | None,
    risk_signal: str | None,
    diagnostics: list[dict[str, Any]],
    opportunities: list[dict[str, Any]],
) -> str:
    trend_text: str | None = None
    for period_label, value in (
        ("同比", summary["final_conversion_yoy"]),
        ("环比", summary["final_conversion_mom"]),
    ):
        if value is None or abs(value) < 0.02:
            continue
        trend_text = (
            f"整体支付转化{period_label}改善"
            if value > 0
            else f"整体支付转化{period_label}承压"
        )
        break
    if trend_text and risk_signal:
        return f"{trend_text}，但{_risk_subject(risk_signal)}仍需关注。"
    if trend_text and positive_signal:
        return f"{trend_text}，{_positive_subject(positive_signal)}是主要正向信号。"
    if trend_text:
        return f"{trend_text}。"
    if risk_signal and positive_signal:
        return (
            f"{_positive_subject(positive_signal)}表现改善，但"
            f"{_risk_subject(risk_signal)}仍需关注。"
        )
    if risk_signal:
        return f"{_risk_subject(risk_signal)}是当前主要经营风险。"
    if positive_signal:
        return f"{_positive_subject(positive_signal)}是当前主要增长信号。"
    if diagnostics:
        return f"{diagnostics[0]['title']}。"
    if opportunities:
        return f"{opportunities[0]['title']}。"
    return "当前可用数据尚未形成明显经营信号。"


def _build_business_evidence(summary: dict[str, Any]) -> list[str]:
    evidence: list[str] = []
    final_rate = summary["final_conversion_rate"]
    final_parts = []
    if final_rate is not None:
        final_parts.append(f"支付转化率 {_percent_text(final_rate)}")
    for period_label, value, unit in (
        (
            "同比",
            summary["final_conversion_yoy"],
            summary["final_conversion_yoy_unit"],
        ),
        (
            "环比",
            summary["final_conversion_mom"],
            summary["final_conversion_mom_unit"],
        ),
    ):
        if value is not None:
            final_parts.append(f"{period_label}{_comparison_text(value, unit)}")
    if final_parts:
        evidence.append("；".join(final_parts))
    decline = summary["largest_declining_stage"]
    if decline is not None:
        period_label = "同比" if decline["comparison"] == "yoy" else "环比"
        evidence.append(
            f"最大拖累：{_business_stage_label(decline['from_label'])}→"
            f"{_business_stage_label(decline['to_label'])}"
            f"{period_label}{_comparison_text(decline['delta'], decline['unit'])}"
        )
    improvement = summary["best_improving_stage"]
    if improvement is not None:
        period_label = (
            "同比" if improvement["comparison"] == "yoy" else "环比"
        )
        evidence.append(
            f"最大改善：{_business_stage_label(improvement['from_label'])}→"
            f"{_business_stage_label(improvement['to_label'])}"
            f"{period_label}{_comparison_text(improvement['delta'], improvement['unit'])}"
        )
    return evidence


def _resolve_business_insight_priority(
    summary: dict[str, Any],
) -> str:
    return summary["diagnosis_level"]


def _business_insight_rank_score(
    summary: dict[str, Any],
    priority: str,
    traffic: int | None,
) -> float:
    priority_score = {
        "stable": 100,
        "improving": 200,
        "attention": 300,
        "high_priority": 400,
    }[priority]
    changes = [
        abs(value)
        for value in (
            summary["final_conversion_yoy"],
            summary["final_conversion_mom"],
            (
                summary["best_improving_stage"]["delta"]
                if summary["best_improving_stage"]
                else None
            ),
            (
                summary["largest_declining_stage"]["delta"]
                if summary["largest_declining_stage"]
                else None
            ),
        )
        if value is not None
    ]
    movement_score = max(changes, default=0) * 100
    scale_score = math.log10(max(traffic or 1, 1))
    return priority_score + movement_score + scale_score


def _risk_subject(signal: str) -> str:
    return re.split(r"同比|环比|当前转化率", signal, maxsplit=1)[0]


def _positive_subject(signal: str) -> str:
    return re.split(r"同比|环比", signal, maxsplit=1)[0]


def _resolve_row_comparison(
    row: pd.Series,
    period: str,
    metric_lookup: dict[str, _SemanticField],
    comparison_lookup: dict[
        tuple[str | None, str | None],
        list[_SemanticField],
    ],
    number_formats: dict[str, tuple[str, ...]],
) -> tuple[float | None, str | None]:
    for target in (
        "payment_conversion_rate",
        "payment_users",
        "traffic_users",
        "gmv",
    ):
        value, unit = _resolve_comparison_for_target(
            row,
            target,
            period,
            metric_lookup.get(target),
            comparison_lookup,
            number_formats,
        )
        if value is not None:
            return value, unit
    generic = comparison_lookup.get((None, period), [])
    if generic:
        comparison = generic[0]
        return (
            _comparison_value(
                row,
                comparison,
                current_metric=None,
                number_formats=number_formats,
            ),
            comparison.comparison_unit,
        )
    return None, None


def _resolve_comparison_for_target(
    row: pd.Series,
    target_metric_key: str,
    period: str,
    current_metric: _SemanticField | None,
    comparison_lookup: dict[
        tuple[str | None, str | None],
        list[_SemanticField],
    ],
    number_formats: dict[str, tuple[str, ...]],
) -> tuple[float | None, str | None]:
    comparisons = comparison_lookup.get((target_metric_key, period), [])
    if not comparisons:
        return None, None
    comparison = comparisons[0]
    value = _comparison_value(
        row,
        comparison,
        current_metric=current_metric,
        number_formats=number_formats,
    )
    return value, (
        "ratio_change"
        if comparison.comparison_value_kind == "baseline" and value is not None
        else comparison.comparison_unit
    )


def _comparison_value(
    row: pd.Series,
    comparison: _SemanticField,
    *,
    current_metric: _SemanticField | None,
    number_formats: dict[str, tuple[str, ...]],
) -> float | None:
    raw_value = row.get(comparison.source_column)
    if comparison.comparison_value_kind == "baseline":
        baseline = _coerce_numeric(raw_value, "absolute", ())
        if baseline is None or baseline == 0 or current_metric is None:
            return None
        current = _metric_value(row, current_metric, number_formats)
        if current is None:
            return None
        return round((current - baseline) / baseline, 6)
    return _coerce_numeric(
        raw_value,
        (
            "percentage_point"
            if comparison.comparison_unit == "percentage_point"
            else "ratio"
            if comparison.comparison_unit == "ratio_change"
            else "absolute"
        ),
        number_formats.get(comparison.source_column, ()),
    )


def _lookup_metric_value(
    row: pd.Series,
    metric_key: str,
    metric_lookup: dict[str, _SemanticField],
    number_formats: dict[str, tuple[str, ...]],
) -> int | float | None:
    metric = metric_lookup.get(metric_key)
    if metric is None:
        return None
    return _metric_value(row, metric, number_formats)


def _metric_value(
    row: pd.Series,
    metric: _SemanticField,
    number_formats: dict[str, tuple[str, ...]],
) -> int | float | None:
    value = _coerce_numeric(
        row.get(metric.source_column),
        metric.unit,
        number_formats.get(metric.source_column, ()),
    )
    if value is None:
        return None
    if metric.unit == "count":
        return int(round(value))
    return round(float(value), 6 if metric.unit == "ratio" else 2)


def _coerce_numeric(
    value: Any,
    unit: str,
    number_formats: Iterable[str],
) -> float | None:
    if not _has_value(value) or isinstance(value, bool):
        return None
    is_percent_text = isinstance(value, str) and "%" in value
    is_percentage_point_text = isinstance(value, str) and "百分点" in value
    numeric = _raw_numeric(value)
    if numeric is None:
        return None
    has_percent_format = any("%" in item for item in number_formats)
    if unit in {"ratio", "percentage_point"}:
        if is_percent_text or is_percentage_point_text:
            numeric /= 100
        elif not has_percent_format and abs(numeric) > 1 and abs(numeric) <= 100:
            numeric /= 100
    return float(numeric)


def _raw_numeric(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if pd.isna(value) or math.isinf(float(value)):
            return None
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        negative = text.startswith("(") and text.endswith(")")
        cleaned = re.sub(
            r"[,%￥¥元人次个+]",
            "",
            text.replace("百分点", ""),
        )
        cleaned = cleaned.strip().strip("()")
        try:
            numeric = float(cleaned)
        except ValueError:
            return None
        return -numeric if negative else numeric
    return None


def _resolve_analysis_status(
    *,
    dimensions: list[_SemanticField],
    business_metrics: list[_SemanticField],
    detail_rows: list[pd.Series],
    recognition_ratio: float,
    warnings: list[str],
) -> str:
    has_multiple_table_warning = any("第二个独立表格" in item for item in warnings)
    if (
        dimensions
        and len(business_metrics) >= 2
        and detail_rows
        and recognition_ratio >= 0.6
        and not has_multiple_table_warning
    ):
        return "ready"
    return "partial"


def _looks_like_second_header(values: list[Any]) -> bool:
    non_empty = [value for value in values if _has_value(value)]
    if len(non_empty) < 2:
        return False
    if any(_raw_numeric(value) is not None for value in non_empty):
        return False
    return sum(_looks_semantic_header(value) for value in non_empty) >= 2


def _looks_semantic_header(value: Any) -> bool:
    normalized = _normalize(str(value))
    hints = (
        "品类",
        "渠道",
        "地区",
        "用户类型",
        "版本",
        "用户数",
        "人数",
        "转化率",
        "gmv",
        "客单价",
        "同比",
        "环比",
    )
    return any(_normalize(hint) in normalized for hint in hints)


def _value_range_bucket(values: list[float]) -> str:
    if not values:
        return "not_numeric"
    minimum, maximum = min(values), max(values)
    if minimum >= -1 and maximum <= 1:
        return "minus_1_to_1"
    if minimum >= 0 and maximum <= 100:
        return "zero_to_100"
    if minimum >= 0 and maximum <= 10000:
        return "zero_to_10000"
    return "wide_range"


def _metric_rule(metric_key: str) -> _MetricRule:
    return next(rule for rule in METRIC_RULES if rule.metric_key == metric_key)


def _rule_to_field(rule: _MetricRule, source_column: str) -> _SemanticField:
    return _SemanticField(
        source_column=source_column,
        label=rule.label,
        semantic_key=rule.metric_key,
        role=rule.role,
        unit=rule.unit,
        aggregation=rule.aggregation,
        confidence="high",
        stage_order=rule.stage_order,
        semantic_role=rule.semantic_role,
    )


def _safe_semantic_key(value: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "_", value.casefold()).strip("_")
    if normalized:
        return normalized
    digest = hashlib.sha1(value.encode("utf-8")).hexdigest()[:8]
    return f"field_{digest}"


def _normalize(value: str) -> str:
    return re.sub(r"[\s_\-—>→/（）()]+", "", str(value).strip().casefold())


def _has_value(value: Any) -> bool:
    if value is None:
        return False
    try:
        if pd.isna(value):
            return False
    except (TypeError, ValueError):
        pass
    return bool(str(value).strip())


def _as_int(value: int | float | None) -> int | None:
    return None if value is None else int(round(value))


def _as_float(value: int | float | None) -> float | None:
    return None if value is None else float(value)


def _percent_text(value: float) -> str:
    return f"{value * 100:.2f}%"


def _signed_percent_text(value: float) -> str:
    return f"{value * 100:+.2f}%"


def _comparison_text(value: float, unit: str | None) -> str:
    if unit == "percentage_point":
        return f"{value * 100:+.2f} 个百分点"
    if unit == "absolute_change":
        return f"{value:+,.2f}"
    return _signed_percent_text(value)


def _optional_percent_text(value: float | None) -> str:
    return "不可用" if value is None else _percent_text(value)


def _business_stage_label(label: str) -> str:
    return label.replace("用户", "").strip()


def _severity_for_change(value: float) -> str:
    magnitude = abs(value)
    if magnitude >= 0.1:
        return "high"
    if magnitude >= 0.05:
        return "medium"
    return "low"


def _is_material_decline(value: float | None, unit: str | None) -> bool:
    if value is None:
        return False
    if unit == "absolute_change":
        return value <= -1
    return value <= -0.05


def _is_severe_decline(value: float, unit: str | None) -> bool:
    if unit == "absolute_change":
        return value <= -10
    return value <= -0.1


def _deduplicate_diagnostics(
    diagnostics: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    seen: set[tuple[str, str | None, str | None]] = set()
    result: list[dict[str, Any]] = []
    for diagnostic in diagnostics:
        key = (
            diagnostic["diagnostic_type"],
            diagnostic.get("dimension_value"),
            diagnostic.get("metric_key"),
        )
        if key in seen:
            continue
        seen.add(key)
        result.append(diagnostic)
    return result


def _rank_and_limit_diagnostics(
    candidates: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for candidate in candidates:
        dimension_value = str(candidate.get("dimension_value") or "__overall__")
        grouped.setdefault(dimension_value, []).append(candidate)

    selected: list[dict[str, Any]] = []
    for dimension_candidates in grouped.values():
        seen_keys: set[str] = set()
        ranked = sorted(
            dimension_candidates,
            key=lambda item: float(item.get("_score", 0)),
            reverse=True,
        )
        for candidate in ranked:
            dedupe_key = str(candidate.get("_dedupe_key", candidate["title"]))
            if dedupe_key in seen_keys:
                continue
            seen_keys.add(dedupe_key)
            selected.append(
                {
                    key: value
                    for key, value in candidate.items()
                    if not key.startswith("_")
                }
            )
            if len(seen_keys) >= 2:
                break
    return selected
