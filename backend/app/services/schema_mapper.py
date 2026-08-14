import json
import logging
from dataclasses import dataclass
from io import BytesIO
from typing import Any, Protocol

import pandas as pd
from openpyxl import load_workbook
from pydantic import ValidationError

from app.core.config import settings
from app.schemas.schema_mapping import SchemaMappingResponse
from app.services.excel_parser import (
    CORE_REQUIRED_COLUMNS,
    ExcelParseError,
    OPTIONAL_FUNNEL_COLUMNS,
    ParsedExcel,
    REQUIRED_COLUMNS,
)


logger = logging.getLogger(__name__)


STANDARD_FIELD_DESCRIPTIONS = {
    "user_id": "用户唯一标识",
    "channel": "用户来源渠道",
    "register_time": "注册时间",
    "view_time": "浏览或访问时间",
    "lead_time": "线索或留资时间",
    "appointment_time": "预约时间",
    "visit_time": "到店或核销时间",
    "pay_time": "支付或成交时间",
    "order_amount": "支付或订单金额",
}


SCHEMA_MAPPING_SYSTEM_PROMPT = (
    f"""你是 GrowthLens AI 的 Excel 字段语义识别器。

任务：把用户上传 Excel 的原始字段映射到以下增长分析标准字段：
{json.dumps(STANDARD_FIELD_DESCRIPTIONS, ensure_ascii=False, indent=2)}

核心必填字段：
{json.dumps(CORE_REQUIRED_COLUMNS, ensure_ascii=False)}

可选漏斗字段：
{json.dumps(OPTIONAL_FUNNEL_COLUMNS, ensure_ascii=False)}

只输出一个合法 JSON 对象，结构必须为：
{{
  "mapping": {{"标准字段": "原始字段或null"}},
  "confidence": {{"标准字段": "high、medium、low或null"}},
  "unmapped_columns": ["未使用的原始字段"]
}}

约束：
1. mapping 的 Key 只能使用上述标准字段。
2. mapping 的 Value 只能逐字使用输入中真实存在的原始字段，
禁止改写或虚构。
3. 同一个原始字段不能映射到多个标准字段。
4. 只有字段名称语义明确时才映射；无法判断时必须返回 null，
不要强制匹配。
5. 不根据行业常识补充输入中不存在的字段。
6. 可选漏斗字段不存在时返回 null，不得为了凑齐字段而强制映射。
7. 不输出 Markdown、代码块、解释或 JSON 之外的内容。
"""
)


AGGREGATE_SEMANTIC_KEYS = {
    "dimension": {
        "category",
        "channel",
        "region",
        "user_type",
        "page_version",
    },
    "count_metric": {
        "traffic_users",
        "product_detail_users",
        "lead_users",
        "appointment_users",
        "sku_selection_users",
        "time_confirmation_users",
        "order_submission_users",
        "visit_users",
        "payment_users",
        "online_payment_users",
        "store_payment_users",
    },
    "rate_metric": {
        "traffic_to_detail_rate",
        "detail_to_appointment_rate",
        "appointment_to_sku_rate",
        "sku_to_time_confirmation_rate",
        "time_confirmation_to_order_rate",
        "order_to_payment_rate",
        "payment_conversion_rate",
    },
    "amount_metric": {
        "gmv",
        "average_order_value",
    },
}


AGGREGATE_SCHEMA_SYSTEM_PROMPT = f"""你是 GrowthLens AI 的聚合经营报表字段识别器。

规则引擎已经处理了语义明确的字段。你只需要判断剩余的低确定性字段，
并且只能使用以下语义目录：
{json.dumps({key: sorted(value) for key, value in AGGREGATE_SEMANTIC_KEYS.items()}, ensure_ascii=False, indent=2)}

只输出合法 JSON：
{{
  "fields": [
    {{
      "source_column": "输入中逐字存在的字段名",
      "role": "dimension | count_metric | rate_metric | amount_metric",
      "semantic_key": "目录中的语义 Key",
      "confidence": "high | medium | low"
    }}
  ]
}}

约束：
1. source_column 只能逐字引用输入中的字段名，不得创造字段。
2. 无法可靠判断的字段不要返回，不能强制映射。
3. 同一字段最多映射一次。
4. 只能依据字段名、数据类型、百分比格式、唯一值比例和值范围分桶判断。
5. 不输出 Markdown、解释或 JSON 之外的内容。
"""


class SchemaMappingError(RuntimeError):
    """Raised when semantic schema mapping cannot be completed."""


class SchemaMappingProvider(Protocol):
    name: str

    def map_columns(self, columns: list[str]) -> dict[str, Any]:
        """Return the provider's raw structured mapping response."""


@dataclass(frozen=True)
class ExtractedExcelSchema:
    sheet_name: str
    header_row_index: int
    columns: tuple[str, ...]
    detected_sheet_names: tuple[str, ...]


@dataclass(frozen=True)
class _SchemaCandidate:
    sheet_name: str
    sheet_index: int
    header_row_index: int
    columns: tuple[str, ...]
    data_row_count: int


def extract_excel_schema(file_content: bytes) -> ExtractedExcelSchema:
    """Select the most likely detail sheet and return its original headers."""
    if not file_content:
        raise ExcelParseError(
            "上传的 Excel 文件为空。",
            error="Excel文件不可用",
        )

    try:
        workbook = load_workbook(
            BytesIO(file_content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise ExcelParseError(
            "Excel 文件无法读取，请确认文件未损坏且格式正确。",
            error="Excel文件不可用",
        ) from exc

    try:
        detected_sheet_names = tuple(str(name) for name in workbook.sheetnames)
        candidates: list[_SchemaCandidate] = []

        for sheet_index, sheet_name in enumerate(detected_sheet_names):
            worksheet = workbook[sheet_name]
            header = _find_header_row(worksheet)
            if header is None:
                continue

            header_row_index, columns = header
            candidates.append(
                _SchemaCandidate(
                    sheet_name=sheet_name,
                    sheet_index=sheet_index,
                    header_row_index=header_row_index,
                    columns=columns,
                    data_row_count=max(worksheet.max_row - header_row_index, 0),
                )
            )

        if not candidates:
            raise ExcelParseError(
                "Excel 文件中未找到可识别的字段行。",
                error="Excel字段不可用",
                detected_sheet_names=list(detected_sheet_names),
            )

        best_candidate = max(
            candidates,
            key=lambda candidate: (
                len(candidate.columns),
                candidate.data_row_count > 0,
                candidate.data_row_count,
                -candidate.header_row_index,
                -candidate.sheet_index,
            ),
        )
        return ExtractedExcelSchema(
            sheet_name=best_candidate.sheet_name,
            header_row_index=best_candidate.header_row_index,
            columns=best_candidate.columns,
            detected_sheet_names=detected_sheet_names,
        )
    finally:
        workbook.close()


def map_columns(
    columns: list[str] | tuple[str, ...],
    *,
    provider: SchemaMappingProvider | None = None,
) -> SchemaMappingResponse:
    """Map uploaded headers and enforce canonical, source-backed results."""
    normalized_columns = _normalize_columns(columns)
    if not normalized_columns:
        return _empty_mapping_response([])

    active_provider = provider or get_schema_mapping_provider()
    raw_result = active_provider.map_columns(normalized_columns)
    return _validate_mapping_result(raw_result, normalized_columns)


def map_aggregate_columns(
    column_profiles: list[dict[str, Any]],
    *,
    client: Any | None = None,
) -> list[dict[str, str]]:
    """Use DeepSeek only as a non-blocking fallback for unresolved headers.

    The payload intentionally contains header names and coarse statistical
    profiles only. Row values and complete business data never leave the
    analysis service.
    """
    profiles = _normalize_aggregate_profiles(column_profiles)
    if not profiles:
        return []
    if settings.ai_provider != "deepseek" or not settings.deepseek_api_key:
        return []

    active_client = client or _create_deepseek_client(
        settings.deepseek_api_key
    )
    try:
        completion = active_client.chat.completions.create(
            model=settings.ai_model,
            messages=[
                {
                    "role": "system",
                    "content": AGGREGATE_SCHEMA_SYSTEM_PROMPT,
                },
                {
                    "role": "user",
                    "content": json.dumps(
                        {"unresolved_columns": profiles},
                        ensure_ascii=False,
                    ),
                },
            ],
            response_format={"type": "json_object"},
            max_tokens=1000,
            stream=False,
        )
        content = completion.choices[0].message.content
        payload = json.loads(content) if content else {}
    except Exception as exc:
        logger.exception("聚合字段 DeepSeek 兜底请求失败：%r", exc)
        return []

    return _validate_aggregate_mapping_payload(payload, profiles)


def build_ai_mapped_excel(
    file_content: bytes,
    extracted_schema: ExtractedExcelSchema,
    mapping_result: SchemaMappingResponse,
) -> ParsedExcel:
    """Load the selected sheet and rename verified AI mappings."""
    missing_core_fields = [
        field
        for field in CORE_REQUIRED_COLUMNS
        if mapping_result.mapping.get(field) is None
    ]
    if missing_core_fields:
        raise ExcelParseError(
            "AI 未能识别全部用户增长分析核心字段。",
            error="AI字段映射不完整",
            missing_fields=missing_core_fields,
            detected_sheet_names=list(extracted_schema.detected_sheet_names),
            candidate_sheet_name=extracted_schema.sheet_name,
            recognized_field_count=sum(
                mapping_result.mapping.get(field) is not None
                for field in REQUIRED_COLUMNS
            ),
        )

    complete_mapping = {
        field: source_field
        for field in REQUIRED_COLUMNS
        if (source_field := mapping_result.mapping.get(field)) is not None
    }
    mapped_source_fields = list(complete_mapping.values())
    duplicated_source_fields = sorted(
        {
            source_field
            for source_field in mapped_source_fields
            if mapped_source_fields.count(source_field) > 1
        }
    )
    if duplicated_source_fields:
        raise ExcelParseError(
            "AI 将同一个 Excel 字段映射到了多个标准字段。",
            error="AI字段映射无效",
            missing_fields=duplicated_source_fields,
            detected_sheet_names=list(extracted_schema.detected_sheet_names),
            candidate_sheet_name=extracted_schema.sheet_name,
            recognized_field_count=0,
        )

    source_columns = set(extracted_schema.columns)
    invalid_source_fields = sorted(
        set(complete_mapping.values()) - source_columns
    )
    if invalid_source_fields:
        raise ExcelParseError(
            "AI 字段映射引用了 Excel 中不存在的字段。",
            error="AI字段映射无效",
            missing_fields=invalid_source_fields,
            detected_sheet_names=list(extracted_schema.detected_sheet_names),
            candidate_sheet_name=extracted_schema.sheet_name,
            recognized_field_count=0,
        )

    try:
        data_frame = pd.read_excel(
            BytesIO(file_content),
            sheet_name=extracted_schema.sheet_name,
            header=extracted_schema.header_row_index - 1,
            engine="openpyxl",
        )
    except Exception as exc:
        raise ExcelParseError(
            "AI 已识别字段，但对应 Excel 数据 Sheet 无法读取。",
            error="Excel工作表不可用",
            detected_sheet_names=list(extracted_schema.detected_sheet_names),
            candidate_sheet_name=extracted_schema.sheet_name,
            recognized_field_count=len(complete_mapping),
        ) from exc

    data_frame.columns = [str(column).strip() for column in data_frame.columns]
    actual_columns = set(data_frame.columns)
    missing_source_columns = sorted(
        set(complete_mapping.values()) - actual_columns
    )
    if missing_source_columns:
        raise ExcelParseError(
            "AI 字段映射引用了 Excel 中不存在的字段。",
            error="AI字段映射无效",
            missing_fields=missing_source_columns,
            detected_sheet_names=list(extracted_schema.detected_sheet_names),
            candidate_sheet_name=extracted_schema.sheet_name,
            recognized_field_count=(
                len(complete_mapping) - len(missing_source_columns)
            ),
        )

    if data_frame.empty:
        raise ExcelParseError(
            f"已识别数据 Sheet“{extracted_schema.sheet_name}”，"
            "但该 Sheet 没有数据行。",
            error="Excel数据为空",
            detected_sheet_names=list(extracted_schema.detected_sheet_names),
            candidate_sheet_name=extracted_schema.sheet_name,
            recognized_field_count=len(complete_mapping),
        )

    rename_map = {
        source_field: standard_field
        for standard_field, source_field in complete_mapping.items()
    }
    return ParsedExcel(
        sheet_name=extracted_schema.sheet_name,
        data_frame=data_frame.rename(columns=rename_map),
        detected_sheet_names=extracted_schema.detected_sheet_names,
        field_mapping=complete_mapping,
    )


def get_schema_mapping_provider() -> SchemaMappingProvider:
    if settings.ai_provider != "deepseek":
        raise SchemaMappingError(
            "字段识别模块当前仅支持 AI_PROVIDER=deepseek。"
        )
    return DeepSeekSchemaMappingProvider(
        api_key=settings.deepseek_api_key,
        model=settings.ai_model,
    )


class DeepSeekSchemaMappingProvider:
    name = "deepseek"

    def __init__(
        self,
        *,
        api_key: str | None,
        model: str,
        client: Any | None = None,
    ) -> None:
        if not api_key:
            raise SchemaMappingError(
                "未配置 DEEPSEEK_API_KEY，无法执行字段识别。"
            )
        if not model:
            raise SchemaMappingError(
                "未配置 AI_MODEL，无法执行字段识别。"
            )

        self.model = model
        self.base_url = "https://api.deepseek.com"
        self._client = client or _create_deepseek_client(api_key)

    def map_columns(self, columns: list[str]) -> dict[str, Any]:
        try:
            completion = self._client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": SCHEMA_MAPPING_SYSTEM_PROMPT},
                    {
                        "role": "user",
                        "content": (
                            "请识别以下 Excel 原始字段：\n"
                            f"{json.dumps(columns, ensure_ascii=False)}"
                        ),
                    },
                ],
                response_format={"type": "json_object"},
                max_tokens=1200,
                stream=False,
                extra_body={"thinking": {"type": "disabled"}},
            )
        except Exception as exc:
            logger.exception("DeepSeek 字段识别请求失败：%r", exc)
            raise SchemaMappingError(
                "DeepSeek 字段识别调用失败，"
                "请检查服务配置或稍后重试。"
            ) from exc

        try:
            content = completion.choices[0].message.content
        except (AttributeError, IndexError, TypeError) as exc:
            raise SchemaMappingError(
                "DeepSeek 字段识别返回了无法解析的响应。"
            ) from exc

        if not content or not content.strip():
            raise SchemaMappingError("DeepSeek 字段识别返回了空结果。")

        try:
            payload = json.loads(content)
        except (json.JSONDecodeError, TypeError) as exc:
            raise SchemaMappingError(
                "DeepSeek 字段识别结果不是合法 JSON。"
            ) from exc

        if not isinstance(payload, dict):
            raise SchemaMappingError(
                "DeepSeek 字段识别结果必须是 JSON 对象。"
            )
        return payload


def _create_deepseek_client(api_key: str) -> Any:
    try:
        from openai import OpenAI
    except ImportError as exc:
        raise SchemaMappingError(
            "缺少 openai SDK，请先安装 backend/requirements.txt。"
        ) from exc

    return OpenAI(
        api_key=api_key,
        base_url="https://api.deepseek.com",
        max_retries=1,
        timeout=30.0,
    )


def _find_header_row(worksheet: Any) -> tuple[int, tuple[str, ...]] | None:
    best_header: tuple[int, tuple[str, ...]] | None = None
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=20, values_only=True),
        start=1,
    ):
        columns = _normalize_columns(row)
        if not columns:
            continue
        if best_header is None or len(columns) > len(best_header[1]):
            best_header = (row_index, tuple(columns))
    return best_header


def _normalize_columns(columns: Any) -> list[str]:
    normalized: list[str] = []
    for value in columns:
        if value is None:
            continue
        column = str(value).strip()
        if not column or column.casefold().startswith("unnamed:"):
            continue
        if column not in normalized:
            normalized.append(column)
    return normalized


def _validate_mapping_result(
    raw_result: dict[str, Any],
    columns: list[str],
) -> SchemaMappingResponse:
    try:
        provider_result = SchemaMappingResponse.model_validate(raw_result)
    except ValidationError as exc:
        raise SchemaMappingError(
            "DeepSeek 字段识别结果不符合约定的 JSON 结构。"
        ) from exc

    column_set = set(columns)
    used_columns: set[str] = set()
    mapping: dict[str, str | None] = {}
    confidence: dict[str, str | None] = {}

    for standard_field in REQUIRED_COLUMNS:
        source_field = provider_result.mapping.get(standard_field)
        field_confidence = provider_result.confidence.get(standard_field)
        if (
            source_field not in column_set
            or source_field in used_columns
            or field_confidence is None
        ):
            mapping[standard_field] = None
            confidence[standard_field] = None
            continue

        mapping[standard_field] = source_field
        confidence[standard_field] = field_confidence
        used_columns.add(source_field)

    return SchemaMappingResponse(
        mapping=mapping,
        confidence=confidence,
        unmapped_columns=[
            column for column in columns if column not in used_columns
        ],
    )


def _empty_mapping_response(columns: list[str]) -> SchemaMappingResponse:
    return SchemaMappingResponse(
        mapping={field: None for field in REQUIRED_COLUMNS},
        confidence={field: None for field in REQUIRED_COLUMNS},
        unmapped_columns=columns,
    )


def _normalize_aggregate_profiles(
    column_profiles: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    seen_columns: set[str] = set()
    allowed_types = {"numeric", "text"}
    allowed_ranges = {
        "not_numeric",
        "minus_1_to_1",
        "zero_to_100",
        "zero_to_10000",
        "wide_range",
    }
    for profile in column_profiles:
        source_column = str(profile.get("source_column", "")).strip()
        if not source_column or source_column in seen_columns:
            continue
        inferred_type = str(profile.get("inferred_type", "text"))
        value_range = str(profile.get("value_range", "not_numeric"))
        normalized.append(
            {
                "source_column": source_column,
                "inferred_type": (
                    inferred_type
                    if inferred_type in allowed_types
                    else "text"
                ),
                "numeric_ratio": _bounded_ratio(
                    profile.get("numeric_ratio")
                ),
                "unique_ratio": _bounded_ratio(
                    profile.get("unique_ratio")
                ),
                "percentage_format": bool(
                    profile.get("percentage_format", False)
                ),
                "value_range": (
                    value_range
                    if value_range in allowed_ranges
                    else "not_numeric"
                ),
            }
        )
        seen_columns.add(source_column)
    return normalized


def _validate_aggregate_mapping_payload(
    payload: Any,
    profiles: list[dict[str, Any]],
) -> list[dict[str, str]]:
    if not isinstance(payload, dict) or not isinstance(
        payload.get("fields"), list
    ):
        return []
    allowed_sources = {
        profile["source_column"] for profile in profiles
    }
    used_sources: set[str] = set()
    results: list[dict[str, str]] = []
    for field in payload["fields"]:
        if not isinstance(field, dict):
            continue
        source_column = field.get("source_column")
        role = field.get("role")
        semantic_key = field.get("semantic_key")
        confidence = field.get("confidence")
        if (
            source_column not in allowed_sources
            or source_column in used_sources
            or role not in AGGREGATE_SEMANTIC_KEYS
            or semantic_key not in AGGREGATE_SEMANTIC_KEYS[role]
            or confidence not in {"high", "medium", "low"}
        ):
            continue
        results.append(
            {
                "source_column": source_column,
                "role": role,
                "semantic_key": semantic_key,
                "confidence": confidence,
            }
        )
        used_sources.add(source_column)
    return results


def _bounded_ratio(value: Any) -> float:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return 0.0
    return round(min(max(numeric, 0.0), 1.0), 2)
