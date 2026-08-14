from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

import app.api.analysis as analysis_api
from app.main import app
from app.schemas.schema_mapping import SchemaMappingResponse
from app.services.analysis_classifier import DEFAULT_ANALYSIS_CONTEXT
from app.services.excel_parser import (
    CORE_REQUIRED_COLUMNS,
    OPTIONAL_FUNNEL_COLUMNS,
    REQUIRED_COLUMNS,
)


client = TestClient(app)
SAMPLE_FILE = (
    Path(__file__).resolve().parents[2]
    / "sample_data"
    / "growthlens_synthetic_user_growth.xlsx"
)


@pytest.fixture(autouse=True)
def stub_analysis_classifier(monkeypatch) -> None:
    monkeypatch.setattr(
        analysis_api,
        "classify_analysis_context",
        lambda _schema_mapping, _columns: (
            DEFAULT_ANALYSIS_CONTEXT.model_copy(deep=True)
        ),
    )


def test_analyze_growth_excel_returns_unified_structure(monkeypatch) -> None:
    def fail_if_ai_mapping_is_called(_columns):
        raise AssertionError("固定模板不应调用 AI 字段识别")

    monkeypatch.setattr(
        analysis_api,
        "map_columns",
        fail_if_ai_mapping_is_called,
    )
    response = client.post(
        "/api/analysis/growth",
        files={
            "file": (
                "growth.xlsx",
                _create_analysis_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert set(payload) == {
        "dataset_type",
        "analysis_status",
        "metadata",
        "data_ingestion",
        "schema_mapping",
        "analysis_context",
        "data_quality",
        "metrics",
        "funnel",
        "channels",
    }
    assert payload["dataset_type"] == "user_level"
    assert payload["analysis_status"] == "ready"
    assert payload["metadata"] == {
        "file_name": "growth.xlsx",
        "data_start_date": "2026-06-01",
        "data_end_date": "2026-06-03",
    }
    assert payload["data_ingestion"]["used_sheet_name"] == "用户增长数据"
    assert payload["data_ingestion"]["recognized_field_count"] == 9
    assert payload["data_ingestion"]["missing_fields"] == []
    assert payload["schema_mapping"] == {
        "mapping": {field: field for field in REQUIRED_COLUMNS},
        "source": "fixed",
        "missing_fields": [],
    }
    assert payload["analysis_context"] == (
        DEFAULT_ANALYSIS_CONTEXT.model_dump(mode="json")
    )
    assert payload["data_quality"]["original_user_count"] == 3
    assert payload["data_quality"]["valid_user_count"] == 3
    assert payload["metrics"]["user_counts"]["registered_users"] == 3
    assert payload["metrics"]["user_counts"]["paid_users"] == 1
    assert payload["metrics"]["conversion_rates"]["view_rate"] == 0.6667
    assert payload["metrics"]["revenue"]["gmv"] == 1599
    assert len(payload["funnel"]["stages"]) == 6
    assert set(payload["channels"]) == {"小红书", "微信", "自然流量"}


def test_synthetic_example_keeps_fixed_mapping_path(monkeypatch) -> None:
    def fail_if_ai_mapping_is_called(_columns):
        raise AssertionError("脱敏示例文件不应调用 AI 字段识别")

    monkeypatch.setattr(
        analysis_api,
        "map_columns",
        fail_if_ai_mapping_is_called,
    )
    response = client.post(
        "/api/analysis/growth",
        files={
            "file": (
                SAMPLE_FILE.name,
                SAMPLE_FILE.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["schema_mapping"]["source"] == "fixed"
    assert payload["data_quality"]["original_user_count"] == 1000


def test_nonstandard_excel_uses_ai_mapping(monkeypatch) -> None:
    custom_columns = (
        "客户编号",
        "获客来源",
        "开户日期",
        "首次浏览",
        "提交线索",
        "预订日期",
        "实际到店",
        "付款日期",
        "实收金额",
    )
    mapping = dict(zip(REQUIRED_COLUMNS, custom_columns, strict=True))

    monkeypatch.setattr(
        analysis_api,
        "map_columns",
        lambda _columns: SchemaMappingResponse(
            mapping=mapping,
            confidence={field: "high" for field in REQUIRED_COLUMNS},
            unmapped_columns=[],
        ),
    )

    response = client.post(
        "/api/analysis/growth",
        files={
            "file": (
                "custom_growth.xlsx",
                _create_analysis_workbook(columns=custom_columns),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["schema_mapping"] == {
        "mapping": mapping,
        "source": "ai",
        "missing_fields": [],
    }
    assert payload["data_ingestion"]["field_mapping"] == mapping
    assert payload["metrics"]["user_counts"]["registered_users"] == 3
    assert payload["metrics"]["user_counts"]["paid_users"] == 1
    assert payload["metrics"]["revenue"]["gmv"] == 1599


def test_ai_mapping_cannot_reference_nonexistent_column(monkeypatch) -> None:
    custom_columns = tuple(f"自定义字段{index}" for index in range(9))
    invalid_mapping = dict(
        zip(REQUIRED_COLUMNS, custom_columns, strict=True)
    )
    invalid_mapping["order_amount"] = "AI虚构金额字段"

    monkeypatch.setattr(
        analysis_api,
        "map_columns",
        lambda _columns: SchemaMappingResponse(
            mapping=invalid_mapping,
            confidence={field: "high" for field in REQUIRED_COLUMNS},
            unmapped_columns=[],
        ),
    )
    monkeypatch.setattr(
        analysis_api,
        "classify_dataset_type",
        lambda _columns: "user_level",
    )
    response = client.post(
        "/api/analysis/growth",
        files={
            "file": (
                "hallucinated_mapping.xlsx",
                _create_analysis_workbook(columns=custom_columns),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 422
    assert response.json()["error"] == "AI字段映射无效"
    assert response.json()["missing_fields"] == ["AI虚构金额字段"]


def test_ai_unrecognized_fields_returns_structured_error(monkeypatch) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "用户明细"
    worksheet.append(list(REQUIRED_COLUMNS[:-2]))
    base_time = datetime(2026, 6, 1, 10, 0)
    worksheet.append(
        [
            "U001",
            "小红书",
            base_time,
            base_time + timedelta(minutes=5),
            base_time + timedelta(minutes=10),
            base_time + timedelta(days=1),
            base_time + timedelta(days=2),
        ]
    )
    file_buffer = BytesIO()
    workbook.save(file_buffer)

    monkeypatch.setattr(
        analysis_api,
        "map_columns",
        lambda _columns: SchemaMappingResponse(
            mapping={field: None for field in REQUIRED_COLUMNS},
            confidence={field: None for field in REQUIRED_COLUMNS},
            unmapped_columns=list(REQUIRED_COLUMNS[:-2]),
        ),
    )

    response = client.post(
        "/api/analysis/growth",
        files={
            "file": (
                "missing_fields.xlsx",
                file_buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["dataset_type"] == "unsupported"
    assert payload["analysis_status"] == "unavailable"
    assert payload["metrics"] is None
    assert payload["funnel"] is None
    assert payload["channels"] is None


def test_aggregate_metrics_enters_placeholder_analysis(monkeypatch) -> None:
    def fail_if_ai_mapping_is_called(_columns):
        raise AssertionError("聚合指标报表不应进入用户字段映射")

    monkeypatch.setattr(
        analysis_api,
        "map_columns",
        fail_if_ai_mapping_is_called,
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "月度经营转化"
    worksheet.append(
        [
            "品类",
            "浏览用户数",
            "商详用户数",
            "预约用户数",
            "支付用户数",
            "支付转化率",
            "同比偏差",
            "环比偏差",
        ]
    )
    worksheet.append(["品类A", 1200, 860, 320, 126, 0.105, -0.01, 0.008])
    file_buffer = BytesIO()
    workbook.save(file_buffer)

    response = client.post(
        "/api/analysis/growth",
        files={
            "file": (
                "monthly_metrics.xlsx",
                file_buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["dataset_type"] == "aggregate_metrics"
    assert payload["analysis_status"] == "unavailable"
    assert payload["metrics"] is None
    assert payload["data_quality"] is None


def test_unsupported_workbook_returns_unavailable_structure(monkeypatch) -> None:
    def fail_if_ai_mapping_is_called(_columns):
        raise AssertionError("不支持的数据结构不应调用用户字段映射")

    monkeypatch.setattr(
        analysis_api,
        "map_columns",
        fail_if_ai_mapping_is_called,
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "说明"
    worksheet.append(["说明", "备注"])
    worksheet.append(["口径说明", "暂无结构化指标"])
    file_buffer = BytesIO()
    workbook.save(file_buffer)

    response = client.post(
        "/api/analysis/growth",
        files={
            "file": (
                "notes.xlsx",
                file_buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["dataset_type"] == "unsupported"
    assert payload["analysis_status"] == "unavailable"
    assert payload["metrics"] is None


def test_core_only_fixed_template_generates_available_funnel(
    monkeypatch,
) -> None:
    def fail_if_ai_mapping_is_called(_columns):
        raise AssertionError("核心字段完整时不应调用 AI 字段识别")

    monkeypatch.setattr(
        analysis_api,
        "map_columns",
        fail_if_ai_mapping_is_called,
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "核心数据"
    worksheet.append(list(CORE_REQUIRED_COLUMNS))
    base_time = datetime(2026, 6, 1, 10, 0)
    worksheet.append(["U001", "小红书", base_time, 1599])
    worksheet.append(["U002", "微信", base_time, None])
    worksheet.append(["U003", "自然流量", base_time, 0])
    file_buffer = BytesIO()
    workbook.save(file_buffer)

    response = client.post(
        "/api/analysis/growth",
        files={
            "file": (
                "core_only.xlsx",
                file_buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["schema_mapping"] == {
        "mapping": {field: field for field in CORE_REQUIRED_COLUMNS},
        "source": "fixed",
        "missing_fields": list(OPTIONAL_FUNNEL_COLUMNS),
    }
    assert payload["data_ingestion"]["missing_fields"] == list(
        OPTIONAL_FUNNEL_COLUMNS
    )
    assert payload["metrics"]["user_counts"]["paid_users"] == 1
    assert payload["metrics"]["conversion_rates"]["paid_rate"] == 0.3333
    assert payload["metrics"]["revenue"]["gmv"] == 1599
    assert [
        stage["key"] for stage in payload["funnel"]["stages"]
    ] == ["registered_users", "paid_users"]


def test_partial_ai_mapping_keeps_optional_funnel_fields_optional(
    monkeypatch,
) -> None:
    columns = (
        "客户编号",
        "获客来源",
        "开户日期",
        "实际到店",
        "实收金额",
    )
    mapping = {
        "user_id": "客户编号",
        "channel": "获客来源",
        "register_time": "开户日期",
        "visit_time": "实际到店",
        "order_amount": "实收金额",
    }
    mapping_response = {field: None for field in REQUIRED_COLUMNS}
    mapping_response.update(mapping)
    confidence = {field: None for field in REQUIRED_COLUMNS}
    confidence.update({field: "high" for field in mapping})
    monkeypatch.setattr(
        analysis_api,
        "map_columns",
        lambda _columns: SchemaMappingResponse(
            mapping=mapping_response,
            confidence=confidence,
            unmapped_columns=[],
        ),
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "自定义用户明细"
    worksheet.append(list(columns))
    base_time = datetime(2026, 6, 1, 10, 0)
    worksheet.append(
        ["U001", "小红书", base_time, base_time + timedelta(days=1), 1999]
    )
    worksheet.append(["U002", "微信", base_time, None, None])
    file_buffer = BytesIO()
    workbook.save(file_buffer)

    response = client.post(
        "/api/analysis/growth",
        files={
            "file": (
                "partial_ai_mapping.xlsx",
                file_buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["schema_mapping"] == {
        "mapping": mapping,
        "source": "ai",
        "missing_fields": [
            "view_time",
            "lead_time",
            "appointment_time",
            "pay_time",
        ],
    }
    assert payload["metrics"]["user_counts"]["paid_users"] == 1
    assert payload["metrics"]["revenue"]["gmv"] == 1999
    assert [
        stage["key"] for stage in payload["funnel"]["stages"]
    ] == ["registered_users", "visit_users", "paid_users"]


def _create_analysis_workbook(
    *,
    columns: tuple[str, ...] = REQUIRED_COLUMNS,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "用户增长数据"
    worksheet.append(list(columns))
    base_time = datetime(2026, 6, 1, 10, 0)

    worksheet.append(
        [
            "U001",
            "小红书",
            base_time,
            base_time + timedelta(minutes=5),
            base_time + timedelta(minutes=10),
            base_time + timedelta(days=1),
            base_time + timedelta(days=2),
            base_time + timedelta(days=2, hours=1),
            1599,
        ]
    )
    worksheet.append(
        [
            "U002",
            "微信",
            base_time,
            base_time + timedelta(minutes=3),
            None,
            None,
            None,
            None,
            None,
        ]
    )
    worksheet.append(
        [
            "U003",
            "自然流量",
            base_time,
            None,
            None,
            None,
            None,
            None,
            None,
        ]
    )

    file_buffer = BytesIO()
    workbook.save(file_buffer)
    return file_buffer.getvalue()
