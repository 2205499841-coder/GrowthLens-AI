from io import BytesIO

import pytest
from openpyxl import Workbook

from app.services.aggregate_analyzer import analyze_aggregate_excel
from app.services.schema_mapper import extract_excel_schema


def _workbook_bytes(
    headers: list[str],
    rows: list[list[object]],
    *,
    percent_columns: tuple[int, ...] = (),
    title_rows: list[list[object]] | None = None,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "经营分析"
    for row in title_rows or []:
        worksheet.append(row)
    worksheet.append(headers)
    header_row = worksheet.max_row
    for row in rows:
        worksheet.append(row)
    for column_index in percent_columns:
        for row_index in range(header_row + 1, worksheet.max_row + 1):
            worksheet.cell(row=row_index, column=column_index).number_format = (
                "0.00%"
            )
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _analyze(content: bytes):
    return analyze_aggregate_excel(
        content,
        file_name="synthetic_metrics.xlsx",
        extracted_schema=extract_excel_schema(content),
        fallback_resolver=lambda _profiles: [],
    )


def _standard_rows() -> list[list[object]]:
    return [
        ["总计", 3000, 2200, 1000, 700, 500, 400, 300, 300000, 1000],
        ["品类A", 2000, 1500, 700, 450, 300, 220, 140, 126000, 900],
        ["品类B", 1000, 700, 300, 250, 200, 180, 160, 174000, 1087.5],
    ]


def _standard_headers() -> list[str]:
    return [
        "品类",
        "浏览用户数",
        "商详用户数",
        "预约用户数",
        "SKU选择用户数",
        "预约时间确认用户数",
        "提交订单用户数",
        "支付用户数",
        "GMV",
        "客单价",
    ]


def test_recognizes_aggregate_category_report() -> None:
    result = _analyze(_workbook_bytes(_standard_headers(), _standard_rows()))

    assert result.dataset_type == "aggregate_metrics"
    assert result.analysis_status == "ready"
    assert result.dataset.grain == ["category"]
    assert result.data_quality.detail_row_count == 2
    assert result.data_quality.total_row_detected is True
    assert {item.metric_key for item in result.kpis} >= {
        "traffic_users",
        "appointment_users",
        "payment_users",
        "payment_conversion_rate",
        "gmv",
        "average_order_value",
    }


def test_recognizes_channel_as_primary_dimension() -> None:
    content = _workbook_bytes(
        ["渠道", "浏览人数", "预约人数", "支付人数", "交易额"],
        [
            ["总计", 2000, 500, 180, 216000],
            ["自然搜索", 1200, 260, 90, 99000],
            ["内容渠道", 800, 240, 90, 117000],
        ],
    )
    result = _analyze(content)

    assert result.analysis_status == "ready"
    assert result.dimensions[0].semantic_key == "channel"
    assert [item.dimension_value for item in result.dimension_performance] == [
        "自然搜索",
        "内容渠道",
    ]


def test_builds_dynamic_funnel_from_available_stages() -> None:
    result = _analyze(_workbook_bytes(_standard_headers(), _standard_rows()))

    assert [stage.metric_key for stage in result.funnel.stages] == [
        "traffic_users",
        "product_detail_users",
        "appointment_users",
        "sku_selection_users",
        "time_confirmation_users",
        "order_submission_users",
        "payment_users",
    ]
    payment_stage = result.funnel.stages[-1]
    assert payment_stage.conversion_rate_from_previous == 0.75
    assert payment_stage.dropoff_count == 100


def test_dynamic_funnel_skips_missing_stages() -> None:
    content = _workbook_bytes(
        ["品类", "浏览用户数", "预约用户数", "支付用户数"],
        [["总计", 1000, 300, 120], ["品类A", 1000, 300, 120]],
    )
    result = _analyze(content)

    assert [stage.metric_key for stage in result.funnel.stages] == [
        "traffic_users",
        "appointment_users",
        "payment_users",
    ]
    assert result.analysis_status == "ready"


def test_missing_gmv_does_not_block_analysis() -> None:
    content = _workbook_bytes(
        ["品类", "浏览用户数", "预约用户数", "支付用户数"],
        [["总计", 1000, 300, 120], ["品类A", 1000, 300, 120]],
    )
    result = _analyze(content)

    assert result.analysis_status == "ready"
    assert "gmv" not in {item.metric_key for item in result.kpis}
    assert result.dimension_performance[0].gmv is None


def test_excel_percentage_format_is_preserved() -> None:
    content = _workbook_bytes(
        ["品类", "支付转化率", "GMV"],
        [["总计", 0.125, 100000], ["品类A", 0.125, 100000]],
        percent_columns=(2,),
    )
    result = _analyze(content)
    rate_kpi = next(
        item for item in result.kpis if item.metric_key == "payment_conversion_rate"
    )

    assert rate_kpi.value == 0.125
    assert result.dimension_performance[0].conversion_rate == 0.125


def test_yoy_ratio_change_is_recognized() -> None:
    content = _workbook_bytes(
        ["品类", "浏览用户数", "支付用户数", "支付转化率同比偏差"],
        [["总计", 1000, 100, -0.04], ["品类A", 1000, 100, -0.04]],
        percent_columns=(4,),
    )
    result = _analyze(content)

    assert result.comparisons[0].period == "yoy"
    assert result.comparisons[0].unit == "percentage_point"
    assert result.dimension_performance[0].yoy == -0.04


def test_mom_percentage_point_change_keeps_distinct_unit() -> None:
    content = _workbook_bytes(
        ["品类", "浏览用户数", "支付用户数", "支付转化率环比偏差（百分点）"],
        [["总计", 1000, 100, -3], ["品类A", 1000, 100, -3]],
    )
    result = _analyze(content)

    assert result.comparisons[0].period == "mom"
    assert result.comparisons[0].comparison_type == "percentage_point_change"
    assert result.dimension_performance[0].mom == -0.03
    assert result.dimension_performance[0].mom_unit == "percentage_point"


def test_overall_conversion_uses_total_numerator_and_denominator() -> None:
    content = _workbook_bytes(
        ["品类", "浏览用户数", "支付用户数", "支付转化率"],
        [
            ["总计", 300, 110, 0.99],
            ["品类A", 100, 10, 0.1],
            ["品类B", 200, 100, 0.5],
        ],
        percent_columns=(4,),
    )
    result = _analyze(content)
    rate_kpi = next(
        item for item in result.kpis if item.metric_key == "payment_conversion_rate"
    )

    assert rate_kpi.value == pytest.approx(110 / 300, abs=1e-6)
    assert rate_kpi.value != pytest.approx((0.1 + 0.5) / 2)
    assert rate_kpi.source == "derived"


def test_total_row_is_excluded_from_dimension_details() -> None:
    result = _analyze(_workbook_bytes(_standard_headers(), _standard_rows()))

    assert result.funnel.scope_dimension_value == "总计"
    assert [item.dimension_value for item in result.dimension_performance] == [
        "品类A",
        "品类B",
    ]


def test_partial_status_keeps_available_results() -> None:
    content = _workbook_bytes(
        ["品类", "浏览用户数", "备注"],
        [["品类A", 1000, "口径待确认"], ["品类B", 600, None]],
    )
    result = _analyze(content)

    assert result.analysis_status == "partial"
    assert result.dataset.analysis_status == "partial"
    assert result.dimension_performance[0].traffic_users == 1000
    assert "备注" in result.data_quality.unrecognized_columns


def test_titles_filters_empty_columns_and_two_level_header() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "月度经营报表"
    worksheet.append(["模拟业务月度经营分析"])
    worksheet.append(["统计周期", "2026-07"])
    worksheet.append([])
    worksheet.append(["品类", "转化指标", None, None, "经营指标", None])
    worksheet.merge_cells("A4:A5")
    worksheet.merge_cells("B4:D4")
    worksheet.merge_cells("E4:F4")
    worksheet.append([None, "浏览用户数", "预约用户数", "支付用户数", "GMV", "客单价"])
    worksheet.append(["总计", 1000, 300, 120, 120000, 1000])
    worksheet.append(["品类A", 1000, 300, 120, 120000, 1000])
    buffer = BytesIO()
    workbook.save(buffer)

    result = _analyze(buffer.getvalue())

    assert result.analysis_status == "ready"
    assert result.dataset.header_rows == [4, 5]
    assert result.dataset.report_period == "2026-07"
    assert result.dataset.filters == {"统计周期": "2026-07"}


def _diagnostic_headers() -> list[str]:
    return [
        "品类",
        "浏览用户数",
        "商详用户数",
        "预约用户数",
        "SKU选择用户数",
        "预约时间确认用户数",
        "提交订单用户数",
        "支付用户数",
        "支付转化率同比偏差（百分点）",
        "支付转化率环比偏差（百分点）",
        "浏览→商详同比偏差（百分点）",
        "商详→预约同比偏差（百分点）",
        "预约→SKU同比偏差（百分点）",
        "SKU→时间确认环比偏差（百分点）",
        "时间确认→提交订单环比偏差（百分点）",
        "提交订单→支付环比偏差（百分点）",
    ]


def test_no_total_row_keeps_overall_kpis_and_funnel_empty() -> None:
    content = _workbook_bytes(
        ["品类", "浏览用户数", "预约用户数", "支付用户数"],
        [
            ["品类A", 1000, 300, 120],
            ["品类B", 800, 260, 100],
        ],
    )
    result = _analyze(content)

    assert result.analysis_status == "ready"
    assert result.data_quality.total_row_detected is False
    assert result.kpis == []
    assert result.funnel.stages == []
    assert len(result.dimension_funnel_diagnostics) == 2


def test_multistage_funnel_yoy_and_mom_are_recognized() -> None:
    content = _workbook_bytes(
        _diagnostic_headers(),
        [
            [
                "证件品类",
                1000,
                800,
                400,
                300,
                250,
                200,
                150,
                0.04,
                0.01,
                0.02,
                -0.0742,
                0.2463,
                0.03,
                0.01,
                0.02,
            ]
        ],
        percent_columns=tuple(range(9, 17)),
    )
    result = _analyze(content)
    summary = result.dimension_funnel_diagnostics[0]

    assert len(summary.stages) == 6
    appointment_to_sku = next(
        stage
        for stage in summary.stages
        if stage.to_metric_key == "sku_selection_users"
    )
    sku_to_time = next(
        stage
        for stage in summary.stages
        if stage.to_metric_key == "time_confirmation_users"
    )
    assert appointment_to_sku.yoy_delta == 0.2463
    assert appointment_to_sku.yoy_unit == "percentage_point"
    assert sku_to_time.mom_delta == 0.03
    assert sku_to_time.mom_unit == "percentage_point"
    assert summary.final_conversion_yoy == 0.04
    assert summary.final_conversion_mom == 0.01


def test_stage_comparison_can_be_derived_from_prior_period_counts() -> None:
    content = _workbook_bytes(
        [
            "品类",
            "浏览用户数",
            "商详用户数",
            "预约用户数",
            "同期浏览用户数",
            "同期商详用户数",
            "同期预约用户数",
            "上期浏览用户数",
            "上期商详用户数",
            "上期预约用户数",
        ],
        [["品类A", 1000, 800, 400, 1000, 700, 420, 1000, 750, 375]],
    )
    result = _analyze(content)
    stages = result.dimension_funnel_diagnostics[0].stages

    assert stages[0].yoy_delta == pytest.approx(0.1)
    assert stages[0].mom_delta == pytest.approx(0.05)
    assert stages[1].yoy_delta == pytest.approx(-0.1)
    assert stages[1].mom_delta == pytest.approx(0.0)


def test_selects_largest_improving_and_declining_stages() -> None:
    content = _workbook_bytes(
        _diagnostic_headers(),
        [
            [
                "证件品类",
                1000,
                800,
                400,
                300,
                250,
                200,
                150,
                0.04,
                0.01,
                0.02,
                -0.0742,
                0.2463,
                0.03,
                0.01,
                0.02,
            ]
        ],
        percent_columns=tuple(range(9, 17)),
    )
    summary = _analyze(content).dimension_funnel_diagnostics[0]

    assert summary.best_improving_stage is not None
    assert summary.best_improving_stage.from_metric_key == "appointment_users"
    assert summary.best_improving_stage.to_metric_key == "sku_selection_users"
    assert summary.best_improving_stage.delta == 0.2463
    assert summary.largest_declining_stage is not None
    assert summary.largest_declining_stage.from_metric_key == "product_detail_users"
    assert summary.largest_declining_stage.to_metric_key == "appointment_users"
    assert summary.largest_declining_stage.delta == -0.0742
    assert summary.diagnosis_level == "high_priority"


def test_strict_scale_conversion_rules_identify_only_clear_outliers() -> None:
    content = _workbook_bytes(
        ["品类", "浏览用户数", "支付用户数"],
        [
            ["高流量低转化", 2000, 100],
            ["常规A", 1000, 150],
            ["常规B", 900, 126],
            ["高转化低规模", 500, 100],
        ],
    )
    result = _analyze(content)
    diagnostic_types = {
        (item.dimension_value, item.diagnostic_type)
        for item in result.diagnostics
    }

    assert (
        "高流量低转化",
        "high_traffic_low_conversion",
    ) in diagnostic_types
    assert (
        "高转化低规模",
        "high_conversion_low_traffic",
    ) in diagnostic_types
    assert all(
        item.dimension_value not in {"常规A", "常规B"}
        for item in result.diagnostics
        if item.diagnostic_type
        in {"high_traffic_low_conversion", "high_conversion_low_traffic"}
    )


def test_scale_rule_does_not_label_moderate_size_as_insufficient() -> None:
    content = _workbook_bytes(
        ["品类", "浏览用户数", "支付用户数"],
        [
            ["品类A", 1200, 144],
            ["品类B", 1000, 140],
            ["品类C", 900, 144],
        ],
    )
    result = _analyze(content)

    assert not any(
        item.diagnostic_type == "high_conversion_low_traffic"
        for item in result.diagnostics
    )


def test_each_dimension_has_at_most_two_nonduplicate_diagnostics() -> None:
    content = _workbook_bytes(
        _diagnostic_headers(),
        [
            [
                "证件品类",
                1000,
                800,
                400,
                300,
                250,
                200,
                150,
                0.04,
                0.01,
                0.02,
                -0.0742,
                0.2463,
                0.03,
                0.01,
                0.02,
            ]
        ],
        percent_columns=tuple(range(9, 17)),
    )
    diagnostics = _analyze(content).diagnostics

    assert len(diagnostics) <= 2
    assert len({item.title for item in diagnostics}) == len(diagnostics)


def test_business_signals_are_merged_into_one_insight_per_dimension() -> None:
    content = _workbook_bytes(
        _diagnostic_headers(),
        [
            [
                "儿童写真",
                1000,
                800,
                400,
                300,
                250,
                200,
                150,
                0.0271,
                0.01,
                0.02,
                -0.0742,
                0.2463,
                0.03,
                0.01,
                0.02,
            ]
        ],
        percent_columns=tuple(range(9, 17)),
    )
    result = _analyze(content)

    assert len(result.business_insights) == 1
    insight = result.business_insights[0]
    assert insight.dimension_value == "儿童写真"
    assert "整体支付转化同比改善" in insight.core_judgement
    assert "商详→预约" in insight.core_judgement
    assert insight.positive_signal is not None
    assert "预约→SKU 选择" in insight.positive_signal
    assert insight.risk_signal is not None
    assert "商详→预约" in insight.risk_signal
    assert any("支付转化率" in item for item in insight.key_evidence)
    assert insight.priority == "high_priority"
    assert len(insight.key_evidence) <= 2
    assert any("最大拖累" in item for item in insight.key_evidence)


def test_business_insights_are_unique_and_limited_to_five() -> None:
    rows = []
    for index in range(7):
        rows.append(
            [
                f"品类{index + 1}",
                1000 + index * 100,
                800,
                400,
                300,
                250,
                200,
                150,
                0.02 + index * 0.01,
                0.01,
                0.02,
                -0.03 - index * 0.005,
                0.04 + index * 0.01,
                0.03,
                0.01,
                0.02,
            ]
        )
    content = _workbook_bytes(
        _diagnostic_headers(),
        rows,
        percent_columns=tuple(range(9, 17)),
    )
    result = _analyze(content)

    assert len(result.business_insights) == 5
    assert len(
        {item.dimension_value for item in result.business_insights}
    ) == 5
    assert result.business_insights[0].dimension_value == "品类7"


def test_payment_outcomes_do_not_create_self_loop() -> None:
    content = _workbook_bytes(
        [
            "品类",
            "浏览用户数",
            "商详用户数",
            "预约用户数",
            "提交订单用户数",
            "支付用户数",
            "线上支付用户数",
            "全域支付用户数",
            "门店签到支付用户数",
            "提交订单→支付同期偏差（百分点）",
        ],
        [
            ["总计", 2000, 1500, 700, 400, 300, 180, 330, 120, -0.03],
            ["品类A", 1200, 900, 420, 250, 170, 100, 185, 70, -0.04],
            ["品类B", 800, 600, 280, 150, 130, 80, 145, 50, 0.02],
        ],
        percent_columns=(10,),
    )
    result = _analyze(content)

    funnel_keys = [stage.metric_key for stage in result.funnel_stages]
    assert funnel_keys.count("payment_users") == 1
    assert "online_payment_users" not in funnel_keys
    assert "all_channel_payment_users" not in funnel_keys
    assert "store_payment_users" not in funnel_keys
    assert len(funnel_keys) == len(set(funnel_keys))
    assert all(
        stage.from_metric_key != stage.to_metric_key
        for summary in result.dimension_funnel_diagnostics
        for stage in summary.stages
    )

    metric_roles = {
        metric.metric_key: metric.semantic_role for metric in result.metrics
    }
    assert metric_roles["payment_users"] == "funnel_stage"
    assert metric_roles["online_payment_users"] == "outcome_metric"
    assert metric_roles["all_channel_payment_users"] == "outcome_metric"
    assert metric_roles["store_payment_users"] == "outcome_metric"
    supplemental_keys = {
        metric.metric_key
        for metric in result.dimension_performance[0].supplemental_outcomes
    }
    assert supplemental_keys == {
        "online_payment_users",
        "all_channel_payment_users",
        "store_payment_users",
    }


def test_payment_stage_comparison_survives_supplemental_payment_columns() -> None:
    content = _workbook_bytes(
        [
            "品类",
            "浏览用户数",
            "提交订单用户数",
            "支付用户数",
            "线上支付用户数",
            "全域支付用户数",
            "提交订单→支付同期偏差（百分点）",
            "提交订单→支付环期偏差（百分点）",
        ],
        [["品类A", 1000, 300, 180, 120, 200, -0.0742, 0.031]],
        percent_columns=(7, 8),
    )
    summary = _analyze(content).dimension_funnel_diagnostics[0]
    payment_edge = next(
        stage
        for stage in summary.stages
        if stage.to_metric_key == "payment_users"
    )

    assert payment_edge.from_metric_key == "order_submission_users"
    assert payment_edge.yoy_delta == -0.0742
    assert payment_edge.mom_delta == 0.031
    assert summary.largest_declining_stage is not None
    assert summary.largest_declining_stage.to_metric_key == "payment_users"
    assert summary.best_improving_stage is not None
    assert summary.best_improving_stage.to_metric_key == "payment_users"


def test_dimension_diagnosis_levels_distinguish_business_states() -> None:
    content = _workbook_bytes(
        ["品类", "浏览用户数", "支付用户数", "支付转化率同比偏差"],
        [
            ["高风险", 2000, 100, -0.12],
            ["关注", 1000, 100, -0.04],
            ["改善", 800, 120, 0.06],
            ["稳定", 600, 90, 0.01],
        ],
        percent_columns=(4,),
    )
    result = _analyze(content)
    levels = {
        item.dimension_value: item.diagnosis_level
        for item in result.dimension_funnel_diagnostics
    }

    assert levels == {
        "高风险": "high_priority",
        "关注": "attention",
        "改善": "improving",
        "稳定": "stable",
    }
