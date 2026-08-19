import json
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

import pandas as pd
import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from pydantic import ValidationError

import app.services.ai_report as ai_report_service
from app.main import app
from app.schemas.ai_report import (
    AIReportRequest,
    AIReportResponse,
    DraftGrowthExplanation,
)
from app.services.aggregate_analyzer import analyze_aggregate_excel
from app.services.ai_report import (
    AIReportProviderError,
    DeepSeekAIReportProvider,
    MockAIReportProvider,
    OpenAIReportProvider,
    build_model_input,
    generate_ai_report,
)
from app.services.data_cleaner import clean_growth_data
from app.services.excel_parser import REQUIRED_COLUMNS
from app.services.growth_metrics import build_growth_analysis
from app.services.schema_mapper import extract_excel_schema


SAMPLE_FILE = (
    Path(__file__).resolve().parents[2]
    / "sample_data"
    / "growthlens_synthetic_user_growth.xlsx"
)
client = TestClient(app)


def test_user_level_endpoint_returns_unified_report(monkeypatch) -> None:
    request = _sample_user_request()
    monkeypatch.setattr(
        "app.services.ai_report.get_ai_report_provider",
        lambda: MockAIReportProvider(),
    )

    response = client.post(
        "/api/ai/report",
        json=request.model_dump(mode="json"),
    )

    assert response.status_code == 200
    report = response.json()
    assert set(report) == {
        "core_conclusion",
        "key_issues",
        "priority_actions",
        "opportunities",
        "limitations",
    }
    assert len(report["key_issues"]) <= 3
    assert len(report["priority_actions"]) <= 3
    assert len(report["opportunities"]) <= 2
    assert all("experiment" not in item for item in report["priority_actions"])
    assert all(
        evidence["evidence_ref"]
        for issue in report["key_issues"]
        for evidence in issue["evidence"]
    )


def test_aggregate_endpoint_returns_unified_report(monkeypatch) -> None:
    request = _sample_aggregate_request()
    monkeypatch.setattr(
        "app.services.ai_report.get_ai_report_provider",
        lambda: MockAIReportProvider(),
    )

    response = client.post(
        "/api/ai/report",
        json=request.model_dump(mode="json"),
    )

    assert response.status_code == 200
    report = response.json()
    assert report["core_conclusion"]
    assert report["growth_explanation"]["dimension_value"] == "品类甲"
    assert report["growth_explanation"]["growth_driver"] == "conversion"
    assert report["growth_explanation"]["evidence"][0]["display_values"]
    assert 1 <= len(report["key_issues"]) <= 3
    assert len(report["priority_actions"]) <= 3
    assert len(report["opportunities"]) <= 2
    assert all(item["experiment"] for item in report["priority_actions"])


def test_ai_report_endpoint_reports_missing_deepseek_key(
    monkeypatch,
    caplog,
) -> None:
    request = _sample_user_request()
    monkeypatch.setattr(
        ai_report_service,
        "settings",
        SimpleNamespace(
            ai_provider="deepseek",
            deepseek_api_key=None,
            openai_api_key=None,
            ai_model="deepseek-v4-pro",
        ),
    )

    with caplog.at_level("ERROR", logger="app.api.ai_reports"):
        response = client.post(
            "/api/ai/report",
            json=request.model_dump(mode="json"),
        )

    assert response.status_code == 503
    assert "DEEPSEEK_API_KEY" in response.json()["detail"]
    assert "AI 报告生成失败" in caplog.text


@pytest.mark.parametrize(
    ("provider_class", "base_url"),
    [
        (DeepSeekAIReportProvider, "https://api.deepseek.com"),
        (OpenAIReportProvider, None),
    ],
)
def test_openai_compatible_provider_uses_json_output(
    provider_class,
    base_url,
) -> None:
    request = _sample_aggregate_request()
    expected_draft = MockAIReportProvider().generate(request)
    expected_report = generate_ai_report(
        request,
        provider=MockAIReportProvider(),
    )
    fake_client = FakeOpenAIClient(expected_draft.model_dump_json())
    provider = provider_class(
        api_key="test-key",
        model="provider-test-model",
        client=fake_client,
    )

    report = generate_ai_report(request, provider=provider)

    assert report == expected_report
    assert provider.base_url == base_url
    call = fake_client.calls[0]
    assert call["model"] == "provider-test-model"
    assert call["response_format"] == {"type": "json_object"}
    assert call["max_tokens"] == 3000
    assert call["stream"] is False
    assert '"dataset_type":"aggregate_metrics"' in call["messages"][1]["content"]
    if provider_class is DeepSeekAIReportProvider:
        assert call["extra_body"] == {
            "thinking": {"type": "disabled"},
        }
    else:
        assert "extra_body" not in call


def test_deepseek_provider_requires_api_key() -> None:
    with pytest.raises(AIReportProviderError, match="DEEPSEEK_API_KEY"):
        DeepSeekAIReportProvider(
            api_key=None,
            model="provider-test-model",
        )


def test_deepseek_provider_logs_api_exception_details(caplog) -> None:
    class FakeResponse:
        status_code = 401
        text = '{"error":"invalid model"}'

    class FakeDeepSeekError(RuntimeError):
        status_code = 401
        response = FakeResponse()

    def raise_deepseek_error(**_kwargs):
        raise FakeDeepSeekError("Unauthorized")

    failing_client = SimpleNamespace(
        chat=SimpleNamespace(
            completions=SimpleNamespace(create=raise_deepseek_error),
        )
    )
    provider = DeepSeekAIReportProvider(
        api_key="test-key",
        model="provider-test-model",
        client=failing_client,
    )

    with caplog.at_level("ERROR", logger="app.services.ai_report"):
        with pytest.raises(AIReportProviderError):
            provider.generate(_sample_user_request())

    assert "DeepSeek API 调用异常" in caplog.text
    assert "status_code=401" in caplog.text
    assert '{"error":"invalid model"}' in caplog.text


@pytest.mark.parametrize("content", ["", "not-json", "{}"])
def test_provider_rejects_invalid_json_or_schema(content: str) -> None:
    provider = DeepSeekAIReportProvider(
        api_key="test-key",
        model="provider-test-model",
        client=FakeOpenAIClient(content),
    )

    with pytest.raises(AIReportProviderError):
        provider.generate(_sample_user_request())


@pytest.mark.parametrize(
    ("field_name", "item_count"),
    [
        ("key_issues", 4),
        ("priority_actions", 4),
        ("opportunities", 3),
    ],
)
def test_unified_report_enforces_output_item_limits(
    field_name: str,
    item_count: int,
) -> None:
    request = _sample_user_request()
    report = generate_ai_report(request, provider=MockAIReportProvider())
    payload = report.model_dump(mode="json")
    source_items = payload[field_name]
    payload[field_name] = [source_items[0] for _ in range(item_count)]

    with pytest.raises(ValidationError):
        AIReportResponse.model_validate(payload)


def test_aggregate_model_input_is_structured_and_excludes_raw_excel() -> None:
    model_input = build_model_input(_sample_aggregate_request())
    payload = json.loads(model_input.split("\n", maxsplit=1)[1])
    context = payload["diagnostic_context"]

    assert context["dataset_type"] == "aggregate_metrics"
    assert {
        "report_period",
        "filters",
        "dimension",
        "dimension_performance",
        "dimension_diagnosis",
        "business_insights",
        "growth_attribution",
        "user_scale_analysis",
        "funnel_contribution_analysis",
        "funnel_summary",
        "detected_anomalies",
        "data_limitations",
    } <= set(context)
    assert payload["evidence_catalog"]
    assert "source_column" not in model_input
    assert "sheet_name" not in model_input
    assert "file_name" not in model_input
    assert "raw_rows" not in model_input
    assert ".xlsx" not in model_input


def test_growth_driver_is_injected_from_backend_conversion() -> None:
    request = _sample_aggregate_request()
    draft = MockAIReportProvider().generate(request)
    report = generate_ai_report(request, provider=StaticProvider(draft))

    assert report.growth_explanation is not None
    assert report.growth_explanation.dimension_value == "品类甲"
    assert report.growth_explanation.growth_driver == "conversion"


def test_growth_driver_is_injected_from_backend_traffic() -> None:
    request = _request_with_first_growth_driver("traffic")
    draft = MockAIReportProvider().generate(request)
    report = generate_ai_report(request, provider=StaticProvider(draft))

    assert report.growth_explanation is not None
    assert report.growth_explanation.dimension_value == "品类甲"
    assert report.growth_explanation.growth_driver == "traffic"


def test_growth_explanation_binds_second_dimension_without_crossing() -> None:
    request = _sample_aggregate_request()
    draft = MockAIReportProvider().generate(request)
    explanation = draft.growth_explanation
    assert explanation is not None
    second_attribution = request.aggregate_analysis.growth_attribution[1]
    stage_reference = ai_report_service._growth_explanation_stage_reference(
        request.aggregate_analysis,
        second_attribution.dimension_value,
    )
    assert stage_reference is not None
    second_explanation = explanation.model_copy(
        update={
            "why": second_attribution.driver_explanation,
            "main_contribution": (
                second_attribution.funnel_contribution_analysis
                .primary_contribution_stage
            ),
            "evidence": [
                explanation.evidence[0].model_copy(
                    update={
                        "evidence_ref": [
                            "aggregate.growth_attribution[1].traffic_change.browse_users_yoy",
                            "aggregate.growth_attribution[1].traffic_change.payment_users_yoy",
                            "aggregate.growth_attribution[1].conversion_change.payment_rate_change",
                            stage_reference,
                        ]
                    }
                )
            ],
        }
    )
    report = generate_ai_report(
        request,
        provider=StaticProvider(
            draft.model_copy(update={"growth_explanation": second_explanation})
        ),
    )

    assert report.growth_explanation is not None
    assert report.growth_explanation.dimension_value == "品类乙"
    assert report.growth_explanation.growth_driver == "combined"
    assert all(
        "品类甲" not in value
        for evidence in report.growth_explanation.evidence
        for value in evidence.display_values
    )


def test_cross_dimension_growth_explanation_degrades_report_only() -> None:
    request = _sample_aggregate_request()
    draft = MockAIReportProvider().generate(request)
    explanation = draft.growth_explanation
    assert explanation is not None
    mixed_evidence = explanation.evidence[0].model_copy(
        update={
            "evidence_ref": [
                "aggregate.growth_attribution[0].traffic_change.browse_users_yoy",
                "aggregate.growth_attribution[1].traffic_change.payment_users_yoy",
            ]
        }
    )
    report = generate_ai_report(
        request,
        provider=StaticProvider(
            draft.model_copy(
                update={
                    "growth_explanation": explanation.model_copy(
                        update={"evidence": [mixed_evidence]}
                    )
                }
            )
        ),
    )

    assert report.growth_explanation is None
    assert "增长来源说明未能与后端归因安全绑定" in report.limitations[-1]
    assert report.key_issues
    assert report.priority_actions


def test_growth_explanation_fabricated_number_degrades_report_only() -> None:
    request = _sample_aggregate_request()
    draft = MockAIReportProvider().generate(request)
    explanation = draft.growth_explanation
    assert explanation is not None
    invalid_explanation = explanation.model_copy(
        update={"why": explanation.why + "预计再提升99%。"}
    )

    report = generate_ai_report(
        request,
        provider=StaticProvider(
            draft.model_copy(
                update={"growth_explanation": invalid_explanation}
            )
        ),
    )

    assert report.growth_explanation is None
    assert "增长来源说明未能与后端归因安全绑定" in report.limitations[-1]
    assert report.key_issues


def test_model_supplied_growth_driver_cannot_override_backend() -> None:
    request = _sample_aggregate_request()
    draft = MockAIReportProvider().generate(request)
    explanation_payload = draft.growth_explanation.model_dump(mode="json")
    explanation_payload["growth_driver"] = "traffic"
    explanation_payload["dimension_value"] = "品类乙"
    attempted_override = DraftGrowthExplanation.model_validate(
        explanation_payload
    )

    report = generate_ai_report(
        request,
        provider=StaticProvider(
            draft.model_copy(
                update={"growth_explanation": attempted_override}
            )
        ),
    )

    assert report.growth_explanation is not None
    assert report.growth_explanation.dimension_value == "品类甲"
    assert report.growth_explanation.growth_driver == "conversion"
    assert not hasattr(attempted_override, "growth_driver")


def test_ai_report_request_rejects_raw_data_fields() -> None:
    payload = _sample_aggregate_request().model_dump(mode="json")
    payload["raw_rows"] = [{"user_id": "U001"}]

    response = client.post("/api/ai/report", json=payload)

    assert response.status_code == 422


def test_report_rejects_unknown_evidence_ref() -> None:
    request = _sample_aggregate_request()
    report = MockAIReportProvider().generate(request)
    evidence = report.key_issues[0].evidence[0].model_copy(
        update={"evidence_ref": ["aggregate.unknown.value"]}
    )
    issue = report.key_issues[0].model_copy(update={"evidence": [evidence]})
    invalid_report = report.model_copy(update={"key_issues": [issue]})

    with pytest.raises(AIReportProviderError, match="未知 evidence_ref"):
        generate_ai_report(
            request,
            provider=StaticProvider(invalid_report),
        )


def test_report_rejects_fabricated_number() -> None:
    request = _sample_aggregate_request()
    report = MockAIReportProvider().generate(request)
    invalid_report = report.model_copy(
        update={"core_conclusion": report.core_conclusion + "预计提升 99%。"}
    )

    with pytest.raises(
        AIReportProviderError,
        match=r"core_conclusion.*99%",
    ):
        generate_ai_report(
            request,
            provider=StaticProvider(invalid_report),
        )


def test_report_rejects_fabricated_number_adjacent_to_chinese_text() -> None:
    request = _sample_aggregate_request()
    draft = MockAIReportProvider().generate(request)
    invalid_draft = draft.model_copy(
        update={"core_conclusion": draft.core_conclusion + "预计提升99%。"}
    )

    with pytest.raises(
        AIReportProviderError,
        match=r"core_conclusion.*99%",
    ):
        generate_ai_report(
            request,
            provider=StaticProvider(invalid_draft),
        )


def test_report_rejects_percentage_point_as_percent() -> None:
    request = _sample_aggregate_request()
    draft = MockAIReportProvider().generate(request)
    issue = draft.key_issues[0]
    evidence = issue.evidence[0]
    invalid_evidence = evidence.model_copy(
        update={"interpretation": "支付转化同比提升 +4%。"}
    )
    invalid_issue = issue.model_copy(update={"evidence": [invalid_evidence]})
    invalid_draft = draft.model_copy(update={"key_issues": [invalid_issue]})

    with pytest.raises(
        AIReportProviderError,
        match=r"evidence\[0\]\.interpretation.*4%",
    ):
        generate_ai_report(
            request,
            provider=StaticProvider(invalid_draft),
        )


def test_report_rejects_changed_decimal_precision() -> None:
    request = _sample_aggregate_request()
    draft = MockAIReportProvider().generate(request)
    invalid_draft = draft.model_copy(
        update={"core_conclusion": draft.core_conclusion + "支付转化率15%。"}
    )

    with pytest.raises(
        AIReportProviderError,
        match=r"core_conclusion.*15%",
    ):
        generate_ai_report(
            request,
            provider=StaticProvider(invalid_draft),
        )


def test_backend_injects_exact_display_values_from_evidence_refs() -> None:
    request = _sample_aggregate_request()

    report = generate_ai_report(request, provider=MockAIReportProvider())
    evidence = report.key_issues[0].evidence[0]

    assert evidence.display_values == [
        "支付转化率 15.00%；同比+4.00 个百分点；环比+1.00 个百分点"
    ]
    assert "15.00" not in evidence.interpretation
    assert "百分点" not in evidence.interpretation


def test_one_evidence_can_hydrate_multiple_refs_in_order() -> None:
    request = _sample_aggregate_request()
    draft = MockAIReportProvider().generate(request)
    issue = draft.key_issues[0]
    combined_evidence = issue.evidence[0].model_copy(
        update={
            "evidence_ref": [
                issue.evidence[0].evidence_ref[0],
                issue.evidence[1].evidence_ref[0],
            ],
            "interpretation": "整体趋势改善，但主要拖累节点仍需验证。",
        }
    )
    combined_issue = issue.model_copy(
        update={"evidence": [combined_evidence]}
    )
    combined_draft = draft.model_copy(
        update={"key_issues": [combined_issue]}
    )

    report = generate_ai_report(
        request,
        provider=StaticProvider(combined_draft),
    )

    assert report.key_issues[0].evidence[0].display_values == [
        "支付转化率 15.00%；同比+4.00 个百分点；环比+1.00 个百分点",
        "最大拖累：商详→预约同比-7.42 个百分点",
    ]


def test_analysis_endpoint_does_not_depend_on_ai_provider(monkeypatch) -> None:
    content = _aggregate_workbook_bytes()

    def fail_if_called():
        raise AssertionError("分析接口不应调用 AI Provider")

    monkeypatch.setattr(
        ai_report_service,
        "get_ai_report_provider",
        fail_if_called,
    )
    response = client.post(
        "/api/analysis/growth",
        files={
            "file": (
                "aggregate_metrics.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert response.json()["dataset_type"] == "aggregate_metrics"


def test_user_level_mock_report_remains_supported() -> None:
    request = _sample_user_request()

    report = generate_ai_report(request, provider=MockAIReportProvider())

    assert report.core_conclusion
    assert report.key_issues
    assert report.priority_actions
    assert report.opportunities


def _sample_user_request() -> AIReportRequest:
    data_frame = pd.read_excel(SAMPLE_FILE, engine="openpyxl")
    analysis = build_growth_analysis(
        clean_growth_data(data_frame),
        file_name=SAMPLE_FILE.name,
    )
    return AIReportRequest.model_validate(
        {
            "dataset_type": "user_level",
            **{
                key: analysis[key]
                for key in (
                    "data_quality",
                    "metrics",
                    "funnel",
                    "channels",
                )
            },
            "schema_mapping": {
                "mapping": {field: field for field in REQUIRED_COLUMNS},
                "source": "fixed",
            },
            "analysis_context": {
                "analysis_type": "user_growth",
                "business_type": "local_service",
                "recommended_metrics": [
                    "注册用户数",
                    "预约率",
                    "到店率",
                    "成交率",
                    "GMV",
                ],
            },
        }
    )


def _sample_aggregate_request() -> AIReportRequest:
    content = _aggregate_workbook_bytes()
    analysis = analyze_aggregate_excel(
        content,
        file_name="synthetic_metrics.xlsx",
        extracted_schema=extract_excel_schema(content),
        fallback_resolver=lambda _profiles: [],
    )
    return AIReportRequest(
        dataset_type="aggregate_metrics",
        aggregate_analysis=analysis,
    )


def _request_with_first_growth_driver(driver: str) -> AIReportRequest:
    request = _sample_aggregate_request()
    analysis = request.aggregate_analysis
    assert analysis is not None
    attributions = list(analysis.growth_attribution)
    attributions[0] = attributions[0].model_copy(
        update={
            "growth_driver": driver,
            "driver_explanation": "支付表现主要伴随流量规模扩大。",
        }
    )
    return request.model_copy(
        update={
            "aggregate_analysis": analysis.model_copy(
                update={"growth_attribution": attributions}
            )
        }
    )


def _aggregate_workbook_bytes() -> bytes:
    headers = [
        "品类",
        "浏览用户数",
        "商详用户数",
        "预约用户数",
        "SKU选择用户数",
        "预约时间确认用户数",
        "提交订单用户数",
        "支付用户数",
        "支付转化率同比偏差（百分点）",
        "支付转化率环比偏差（百分点）",
        "商详→预约同比偏差（百分点）",
        "预约→SKU同比偏差（百分点）",
        "同期浏览用户数",
        "同期预约用户数",
        "同期支付用户数",
    ]
    rows = [
        [
            "品类甲", 1000, 800, 400, 300, 250, 200, 150,
            0.04, 0.01, -0.0742, 0.2463, 1000, 380, 120,
        ],
        [
            "品类乙", 600, 500, 300, 240, 210, 180, 150,
            0.02, 0.01, -0.02, 0.05, 550, 270, 130,
        ],
    ]
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "经营分析"
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    for column_index in range(9, 13):
        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_index, column=column_index).number_format = "0.00%"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class FakeOpenAIClient:
    def __init__(self, content: str) -> None:
        self.calls: list[dict] = []
        self.chat = SimpleNamespace(
            completions=SimpleNamespace(create=self._create),
        )
        self._content = content

    def _create(self, **kwargs):
        self.calls.append(kwargs)
        return SimpleNamespace(
            choices=[
                SimpleNamespace(
                    message=SimpleNamespace(content=self._content),
                )
            ]
        )


class StaticProvider:
    name = "static"

    def __init__(self, report) -> None:
        self.report = report

    def generate(self, _request):
        return self.report
