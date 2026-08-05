from datetime import datetime
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.main import app
from app.services.excel_parser import (
    CORE_REQUIRED_COLUMNS,
    OPTIONAL_FUNNEL_COLUMNS,
    REQUIRED_COLUMNS,
)


client = TestClient(app)

CHINESE_COLUMNS = (
    "用户ID",
    "来源渠道",
    "注册日期",
    "浏览时间",
    "咨询时间",
    "预约时间",
    "到店时间",
    "成交时间",
    "支付金额",
)


def append_growth_row(
    worksheet,
    *,
    columns: tuple[str, ...] = REQUIRED_COLUMNS,
) -> None:
    values = {
        "user_id": "U10001",
        "channel": "小红书",
        "register_time": datetime(2026, 7, 1, 10, 0),
        "view_time": datetime(2026, 7, 1, 10, 5),
        "lead_time": datetime(2026, 7, 1, 10, 10),
        "appointment_time": datetime(2026, 7, 2, 14, 0),
        "visit_time": datetime(2026, 7, 5, 11, 0),
        "pay_time": datetime(2026, 7, 5, 13, 0),
        "order_amount": 1299,
    }
    worksheet.append([values.get(column) for column in columns])


def create_workbook_bytes(
    *,
    columns: tuple[str, ...] = REQUIRED_COLUMNS,
    include_data_row: bool = True,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "用户增长数据"
    worksheet.append(list(columns))

    if include_data_row:
        append_growth_row(worksheet, columns=columns)

    file_buffer = BytesIO()
    workbook.save(file_buffer)
    return file_buffer.getvalue()


def test_parse_valid_excel() -> None:
    response = client.post(
        "/api/uploads/parse",
        headers={"Origin": "http://127.0.0.1:3000"},
        files={
            "file": (
                "photo_growth.xlsx",
                create_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["file_name"] == "photo_growth.xlsx"
    assert payload["sheet_name"] == "用户增长数据"
    assert payload["data_ingestion"] == {
        "used_sheet_name": "用户增长数据",
        "detected_sheet_names": ["用户增长数据"],
        "recognized_field_count": len(REQUIRED_COLUMNS),
        "total_required_field_count": len(REQUIRED_COLUMNS),
        "missing_fields": [],
        "row_count": 1,
        "data_quality_status": "ready",
        "field_mapping": {column: column for column in REQUIRED_COLUMNS},
    }
    assert payload["row_count"] == 1
    assert payload["column_count"] == len(REQUIRED_COLUMNS)
    assert payload["preview"][0]["user_id"] == "U10001"
    assert payload["preview"][0]["order_amount"] == 1299
    assert (
        response.headers["access-control-allow-origin"]
        == "http://127.0.0.1:3000"
    )


def test_reject_non_xlsx_file() -> None:
    response = client.post(
        "/api/uploads/parse",
        files={"file": ("users.csv", b"user_id,channel", "text/csv")},
    )

    assert response.status_code == 422
    assert response.json()["error"] == "Excel文件格式不支持"
    assert response.json()["message"] == "当前仅支持 .xlsx 格式的 Excel 文件。"


def test_return_structured_error_for_missing_fields() -> None:
    response = client.post(
        "/api/uploads/parse",
        files={
            "file": (
                "missing_columns.xlsx",
                create_workbook_bytes(columns=REQUIRED_COLUMNS[:-2]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 422
    assert response.json() == {
        "error": "Excel字段不完整",
        "message": "未找到用户增长分析所需核心字段",
        "missing_fields": ["order_amount"],
        "detected_sheet_names": ["用户增长数据"],
        "candidate_sheet_name": "用户增长数据",
        "recognized_field_count": 7,
        "data_quality_status": "invalid",
    }


def test_parse_accepts_core_fields_without_optional_funnel_fields() -> None:
    response = client.post(
        "/api/uploads/parse",
        files={
            "file": (
                "core_fields.xlsx",
                create_workbook_bytes(columns=CORE_REQUIRED_COLUMNS),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["data_ingestion"]["recognized_field_count"] == len(
        CORE_REQUIRED_COLUMNS
    )
    assert payload["data_ingestion"]["missing_fields"] == list(
        OPTIONAL_FUNNEL_COLUMNS
    )
    assert payload["data_ingestion"]["field_mapping"] == {
        field: field for field in CORE_REQUIRED_COLUMNS
    }


def test_reject_workbook_without_data_rows() -> None:
    response = client.post(
        "/api/uploads/parse",
        files={
            "file": (
                "empty.xlsx",
                create_workbook_bytes(include_data_row=False),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 422
    assert response.json()["error"] == "Excel数据为空"
    assert response.json()["candidate_sheet_name"] == "用户增长数据"


def test_select_detail_sheet_when_overview_sheet_is_first() -> None:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "数据概览"
    overview.append(["指标", "结果"])
    overview.append(["注册用户", 1])

    detail = workbook.create_sheet("用户明细")
    detail.append(list(REQUIRED_COLUMNS))
    append_growth_row(detail)

    file_buffer = BytesIO()
    workbook.save(file_buffer)
    response = client.post(
        "/api/uploads/parse",
        files={
            "file": (
                "overview_first.xlsx",
                file_buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["sheet_name"] == "用户明细"
    assert payload["data_ingestion"]["used_sheet_name"] == "用户明细"
    assert payload["data_ingestion"]["detected_sheet_names"] == [
        "数据概览",
        "用户明细",
    ]
    assert payload["data_ingestion"]["recognized_field_count"] == 9


def test_map_chinese_field_aliases() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "中文用户明细"
    worksheet.append(list(CHINESE_COLUMNS))

    canonical_values = {
        "user_id": "CN10001",
        "channel": "微信",
        "register_time": datetime(2026, 7, 1, 10, 0),
        "view_time": datetime(2026, 7, 1, 10, 5),
        "lead_time": datetime(2026, 7, 1, 10, 10),
        "appointment_time": datetime(2026, 7, 2, 14, 0),
        "visit_time": datetime(2026, 7, 5, 11, 0),
        "pay_time": datetime(2026, 7, 5, 13, 0),
        "order_amount": 1999,
    }
    worksheet.append([canonical_values[column] for column in REQUIRED_COLUMNS])

    file_buffer = BytesIO()
    workbook.save(file_buffer)
    response = client.post(
        "/api/uploads/parse",
        files={
            "file": (
                "chinese_columns.xlsx",
                file_buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["preview"][0]["user_id"] == "CN10001"
    assert payload["preview"][0]["order_amount"] == 1999
    assert payload["data_ingestion"]["field_mapping"] == dict(
        zip(REQUIRED_COLUMNS, CHINESE_COLUMNS, strict=True)
    )
